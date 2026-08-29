"""Offline tests for the edit_excel tool.

No DB: uses the in-memory fallback file store as the file source, exactly like
`tests/test_excel_read_tools.py`. Builds a real .xlsx on disk, drives the tool
fn, then re-opens both the original and the produced file from disk.
"""

from __future__ import annotations

import asyncio
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from app.files.store import CSV_MEDIA_TYPE, XLSX_MEDIA_TYPE, file_store
from app.tools.local import edit_excel


@pytest.fixture(autouse=True)
def _configure_store(tmp_path):
    file_store.configure(str(tmp_path))
    yield


def _make_xlsx(rows: list[list], filename="book.xlsx", title="Data"):
    wb = Workbook()
    ws = wb.active
    ws.title = title
    for row in rows:
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    return asyncio.run(
        file_store.save(buf.getvalue(), filename=filename, media_type=XLSX_MEDIA_TYPE)
    )


def _edit(args):
    return asyncio.run(edit_excel.SPEC.func(args))


def _new_id(out: str) -> str:
    """Pull the produced file id out of the tool's download line."""
    marker = "GET /v1/files/"
    assert marker in out, out
    return out.split(marker, 1)[1].split()[0].strip().rstrip(".")


GRID = [["item", "qty"], ["a", 1], ["b", 2], ["c", 3]]


def test_the_edit_lands_in_a_new_file_and_the_original_is_untouched():
    rec = _make_xlsx(GRID)
    before = Path(rec.path).read_bytes()

    out = _edit({"file_id": rec.id, "operation": "set_cells",
                 "cells": [{"cell": "B2", "value": 99}]})

    assert Path(rec.path).read_bytes() == before, "the uploaded file was modified"
    edited = load_workbook(file_store.get(_new_id(out)).path)
    assert edited["Data"]["B2"].value == 99


def test_a_formula_elsewhere_in_the_workbook_survives_the_edit():
    """The data_only regression guard. Loading data_only=True and saving replaces
    every formula with its cached value (or nothing), silently turning a live
    spreadsheet into static numbers. Measured on openpyxl 3.1.5."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["item", "qty", "total"])
    ws.append(["a", 2, "=B2*10"])
    buf = BytesIO()
    wb.save(buf)
    rec = asyncio.run(
        file_store.save(buf.getvalue(), filename="f.xlsx", media_type=XLSX_MEDIA_TYPE)
    )

    out = _edit({"file_id": rec.id, "operation": "set_cells",
                 "cells": [{"cell": "B2", "value": 5}]})

    edited = load_workbook(file_store.get(_new_id(out)).path)
    assert edited["Data"]["C2"].value == "=B2*10"


def test_an_unknown_file_id_is_an_error():
    out = _edit({"file_id": "nope", "operation": "set_cells",
                 "cells": [{"cell": "A1", "value": 1}]})
    assert out.startswith("ERROR:")
    assert "no such file" in out


def test_a_csv_is_refused_with_a_reason():
    rec = asyncio.run(
        file_store.save(b"a,b\n1,2\n", filename="t.csv", media_type=CSV_MEDIA_TYPE)
    )
    out = _edit({"file_id": rec.id, "operation": "append_rows", "rows": [["x", "y"]]})
    assert out.startswith("ERROR:")
    assert ".csv" in out


def test_an_unknown_operation_is_an_error_naming_the_valid_ones():
    rec = _make_xlsx(GRID)
    out = _edit({"file_id": rec.id, "operation": "sort_rows"})
    assert out.startswith("ERROR:")
    assert "set_cells" in out


def test_a_refused_edit_produces_no_file_at_all():
    """A rejected operation must not leave a half-edited artifact behind."""
    rec = _make_xlsx(GRID)
    before = set(file_store._records)
    out = _edit({"file_id": rec.id, "operation": "set_cells",
                 "cells": [{"cell": "B2", "value": "=1+1"}]})
    assert out.startswith("ERROR:")
    assert set(file_store._records) == before


def test_the_result_names_the_sheet_and_what_changed():
    rec = _make_xlsx(GRID)
    out = _edit({"file_id": rec.id, "operation": "delete_rows", "rows": [2]})
    assert "Data" in out and "deleted 1 data row" in out


def test_it_edits_the_sheet_the_caller_named():
    wb = Workbook()
    wb.remove(wb.active)
    for name in ("First", "Second"):
        wb.create_sheet(name).append(["h"])
    buf = BytesIO()
    wb.save(buf)
    rec = asyncio.run(
        file_store.save(buf.getvalue(), filename="m.xlsx", media_type=XLSX_MEDIA_TYPE)
    )
    out = _edit({"file_id": rec.id, "sheet": "Second", "operation": "append_rows",
                 "rows": [["v"]]})
    edited = load_workbook(file_store.get(_new_id(out)).path)
    assert edited["Second"]["A2"].value == "v"
    assert edited["First"].max_row == 1


def test_a_workbook_over_the_cell_cap_is_refused(monkeypatch):
    """The cap is a module constant, not a tool argument: the model must not be
    able to raise the limit that protects the process."""
    monkeypatch.setattr(edit_excel, "MAX_EDIT_CELLS", 3)
    rec = _make_xlsx(GRID)
    out = _edit({"file_id": rec.id, "operation": "set_cells",
                 "cells": [{"cell": "B2", "value": 1}]})
    assert out.startswith("ERROR:")
    assert "too large" in out


def test_the_cap_is_not_a_tool_argument():
    assert "max_cells" not in str(edit_excel.SPEC.parameters)


# --------------------------------------------------------------------------- #
# Code-review findings (2026-08-29)
# --------------------------------------------------------------------------- #
@pytest.mark.parametrize(
    "operation", ["set_cells", "append_rows", "delete_rows", "delete_columns"]
)
def test_a_missing_operand_is_a_clean_error_not_a_python_traceback(operation):
    """Dropping absent keys let the operand default to a missing positional
    argument, so the model got '_append_rows() missing 1 required positional
    argument' — an internal function name instead of what to supply."""
    rec = _make_xlsx(GRID)
    out = _edit({"file_id": rec.id, "operation": operation})
    assert out.startswith("ERROR:")
    assert "positional argument" not in out
    assert "_" + operation not in out


def test_a_non_spreadsheet_upload_names_its_own_extension():
    """Every non-.xlsx file was reported as '.csv' and routed to read_excel,
    which then fails on a PDF and sends the model round a loop."""
    rec = asyncio.run(
        file_store.save(b"%PDF-1.4\n", filename="report.pdf", media_type="application/pdf")
    )
    out = _edit({"file_id": rec.id, "operation": "append_rows", "rows": [["x"]]})
    assert out.startswith("ERROR:")
    assert ".pdf" in out
    assert ".csv" not in out
    assert "read_document" in out
