"""Offline tests for the pure spreadsheet editor (`app/files/editor.py`).

No DB, no file store — builds openpyxl workbooks in memory and applies one
operation to them. The tool layer that resolves a file_id and saves a NEW file
is tested separately in `tests/test_edit_excel_tool.py`.
"""

from __future__ import annotations

import pytest
from openpyxl import Workbook

from app.files import editor


def _book(rows: list[list], title: str = "Data") -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = title
    for row in rows:
        ws.append(row)
    return wb


GRID = [["item", "qty"], ["a", 1], ["b", 2], ["c", 3]]


# --------------------------------------------------------------------------- #
# set_cells
# --------------------------------------------------------------------------- #
def test_set_cells_writes_the_value_into_the_named_cell():
    wb = _book(GRID)
    editor.apply(wb, operation="set_cells", cells=[{"cell": "B2", "value": 99}])
    assert wb["Data"]["B2"].value == 99


def test_set_cells_refuses_a_formula_value():
    wb = _book(GRID)
    with pytest.raises(editor.EditError, match="formula"):
        editor.apply(wb, operation="set_cells", cells=[{"cell": "B2", "value": "=SUM(B2:B4)"}])


def test_a_refused_formula_leaves_every_cell_in_the_batch_unchanged():
    """A batch is all-or-nothing: a bad cell late in the list must not leave
    the earlier ones written, or the model cannot tell what state the file is in."""
    wb = _book(GRID)
    with pytest.raises(editor.EditError):
        editor.apply(
            wb,
            operation="set_cells",
            cells=[{"cell": "B2", "value": 99}, {"cell": "B3", "value": "=1+1"}],
        )
    assert wb["Data"]["B2"].value == 1


def test_set_cells_refuses_a_malformed_cell_reference():
    wb = _book(GRID)
    with pytest.raises(editor.EditError, match="cell reference"):
        editor.apply(wb, operation="set_cells", cells=[{"cell": "B2:C3", "value": 1}])


# --------------------------------------------------------------------------- #
# append_rows
# --------------------------------------------------------------------------- #
def test_append_rows_adds_after_the_last_data_row():
    wb = _book(GRID)
    editor.apply(wb, operation="append_rows", rows=[["d", 4], ["e", 5]])
    ws = wb["Data"]
    assert ws.max_row == 6
    assert [c.value for c in ws[5]] == ["d", 4]
    assert [c.value for c in ws[6]] == ["e", 5]


def test_append_rows_refuses_a_formula_in_any_cell():
    wb = _book(GRID)
    with pytest.raises(editor.EditError, match="formula"):
        editor.apply(wb, operation="append_rows", rows=[["d", "=1+1"]])
    assert wb["Data"].max_row == 4


# --------------------------------------------------------------------------- #
# add_column
# --------------------------------------------------------------------------- #
def test_add_column_appends_a_header_and_its_values():
    wb = _book(GRID)
    editor.apply(wb, operation="add_column", header="note", values=["x", "y", "z"])
    ws = wb["Data"]
    assert [c.value for c in ws["C"]] == ["note", "x", "y", "z"]


def test_add_column_refuses_a_value_count_that_does_not_match_the_data_rows():
    """Silently padding or dropping would give the model a file it cannot
    reason about; say how many rows there are instead."""
    wb = _book(GRID)
    with pytest.raises(editor.EditError, match="3 data row"):
        editor.apply(wb, operation="add_column", header="note", values=["x"])


# --------------------------------------------------------------------------- #
# delete_rows
# --------------------------------------------------------------------------- #
def test_delete_rows_removes_the_named_data_row():
    wb = _book(GRID)
    editor.apply(wb, operation="delete_rows", rows=[2])
    ws = wb["Data"]
    assert [[c.value for c in r] for r in ws.iter_rows()] == [
        ["item", "qty"], ["a", 1], ["c", 3],
    ]


def test_delete_rows_resolves_every_number_against_the_original_numbering():
    """The model names rows as read_excel displayed them. Deleting 1 then 3 must
    remove the ORIGINAL rows 1 and 3 — applying them top-down would shift the
    second target and silently delete the wrong record."""
    wb = _book([["item"], ["a"], ["b"], ["c"], ["d"]])
    editor.apply(wb, operation="delete_rows", rows=[1, 3])
    ws = wb["Data"]
    assert [c[0].value for c in ws.iter_rows()] == ["item", "b", "d"]


def test_delete_rows_refuses_a_row_number_past_the_end():
    wb = _book(GRID)
    with pytest.raises(editor.EditError, match="3 data row"):
        editor.apply(wb, operation="delete_rows", rows=[9])
    assert wb["Data"].max_row == 4


def test_delete_rows_refuses_the_header_row():
    """Row numbers are 1-based into DATA rows, so 0 is not addressable and the
    header can never be removed by this operation."""
    wb = _book(GRID)
    with pytest.raises(editor.EditError, match="out of range"):
        editor.apply(wb, operation="delete_rows", rows=[0])


# --------------------------------------------------------------------------- #
# delete_columns
# --------------------------------------------------------------------------- #
def test_delete_columns_removes_a_column_by_its_header_name():
    wb = _book(GRID)
    editor.apply(wb, operation="delete_columns", columns=["qty"])
    ws = wb["Data"]
    assert [[c.value for c in r] for r in ws.iter_rows()] == [["item"], ["a"], ["b"], ["c"]]


def test_delete_columns_refuses_an_unknown_header():
    wb = _book(GRID)
    with pytest.raises(editor.EditError, match="no column"):
        editor.apply(wb, operation="delete_columns", columns=["nope"])
    assert wb["Data"].max_column == 2


def test_a_large_batch_summarises_the_cell_list_instead_of_naming_every_one():
    """The result goes back into the model's context, where agent/loop.py cuts
    it from the END at MAX_TOOL_RESULT_CHARS. Naming 400 refs would spend the
    budget on a list nobody reads."""
    wb = _book(GRID)
    cells = [{"cell": f"D{i}", "value": i} for i in range(1, 41)]
    result = editor.apply(wb, operation="set_cells", cells=cells)
    assert "set 40 cell(s)" in result.detail
    assert "and 30 more" in result.detail
    assert len(result.detail) < 200


# --------------------------------------------------------------------------- #
# Code-review findings (2026-08-29)
# --------------------------------------------------------------------------- #
def test_two_names_for_the_same_column_delete_it_once():
    """Dedup must be on the RESOLVED index, not the caller's string: 'qty' and
    'B' name one column, and deleting twice destroys the NEXT column instead."""
    wb = _book([["item", "qty", "price"], ["a", 1, 10]])
    editor.apply(wb, operation="delete_columns", columns=["qty", "B"])
    assert [c.value for c in wb["Data"][1]] == ["item", "price"]


def test_a_case_and_space_variant_of_one_header_deletes_it_once():
    wb = _book([["item", "qty", "price"], ["a", 1, 10]])
    editor.apply(wb, operation="delete_columns", columns=["qty", "QTY "])
    assert [c.value for c in wb["Data"][1]] == ["item", "price"]


def test_delete_rows_refuses_a_sheet_that_contains_a_formula():
    """openpyxl is NOT formula-aware: deleting a row leaves '=SUM(B2:B4)' spelled
    exactly as it was, now pointing at the wrong records. Refuse rather than
    hand back a workbook whose formulas silently mean something else."""
    wb = _book(GRID)
    wb["Data"]["B5"] = "=SUM(B2:B4)"
    with pytest.raises(editor.EditError, match="formula"):
        editor.apply(wb, operation="delete_rows", rows=[1])
    assert wb["Data"]["B5"].value == "=SUM(B2:B4)"
    assert wb["Data"].max_row == 5


def test_delete_columns_refuses_a_sheet_that_contains_a_formula():
    wb = _book(GRID)
    wb["Data"]["C2"] = "=B2*2"
    with pytest.raises(editor.EditError, match="formula"):
        editor.apply(wb, operation="delete_columns", columns=["qty"])


def test_delete_rows_refuses_when_another_sheet_has_a_formula_pointing_here():
    """A cross-sheet reference breaks the same way and is invisible from the
    target sheet alone."""
    wb = _book(GRID)
    other = wb.create_sheet("Summary")
    other["A1"] = "=SUM(Data!B2:B4)"
    with pytest.raises(editor.EditError, match="formula"):
        editor.apply(wb, operation="delete_rows", sheet="Data", rows=[1])


def test_deletes_still_work_on_a_formula_free_sheet():
    """The refusal must not make the operation useless on the ordinary case."""
    wb = _book(GRID)
    wb.create_sheet("Notes")["A1"] = "just text"
    editor.apply(wb, operation="delete_rows", rows=[1])
    assert wb["Data"].max_row == 3


def test_set_cells_refuses_a_column_past_excels_last_column():
    """'ZZZ' parses as column 18278; Excel stops at XFD (16384). openpyxl writes
    it without complaint, and the bloated dimension then trips the cell cap on
    every later edit of that file."""
    wb = _book(GRID)
    with pytest.raises(editor.EditError, match="cell reference"):
        editor.apply(wb, operation="set_cells", cells=[{"cell": "ZZZ1", "value": 1}])


def test_set_cells_refuses_a_row_past_excels_last_row():
    wb = _book(GRID)
    with pytest.raises(editor.EditError, match="cell reference"):
        editor.apply(wb, operation="set_cells", cells=[{"cell": "A1048577", "value": 1}])


def test_set_cells_accepts_the_last_legal_cell():
    wb = _book(GRID)
    editor.apply(wb, operation="set_cells", cells=[{"cell": "XFD1048576", "value": 1}])
    assert wb["Data"]["XFD1048576"].value == 1


def test_a_cells_entry_without_a_value_key_is_refused_not_blanked():
    """Nothing enforces the JSON schema's `required`, so an omitted 'value'
    arrived as None and silently cleared the cell while reporting success."""
    wb = _book(GRID)
    with pytest.raises(editor.EditError, match="value"):
        editor.apply(wb, operation="set_cells", cells=[{"cell": "B2"}])
    assert wb["Data"]["B2"].value == 1


def test_an_explicit_null_value_still_clears_the_cell():
    """Refusing an ABSENT key must not take away the ability to blank a cell."""
    wb = _book(GRID)
    editor.apply(wb, operation="set_cells", cells=[{"cell": "B2", "value": None}])
    assert wb["Data"]["B2"].value is None


def test_add_column_on_an_empty_sheet_starts_at_column_a():
    """openpyxl reports max_column == 1 for a BLANK sheet, so 'max_column + 1'
    silently skips column A."""
    from openpyxl import Workbook

    wb = Workbook()
    wb.active.title = "Data"
    editor.apply(wb, operation="add_column", header="h", values=[])
    assert wb["Data"]["A1"].value == "h"


def test_the_formula_refusal_does_not_send_the_model_to_create_excel():
    """create_excel from a capped read_excel dump silently drops rows — the very
    failure read_excel's routing cross-reference warns about. The refusal must
    not recommend it."""
    wb = _book(GRID)
    wb["Data"]["B5"] = "=SUM(B2:B4)"
    with pytest.raises(editor.EditError) as exc:
        editor.apply(wb, operation="delete_rows", rows=[1])
    assert "set_cells" in str(exc.value)
    assert "Do NOT rebuild the file with create_excel" in str(exc.value)
