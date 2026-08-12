"""Parsing for formats that need no heavy dependency: xlsx, csv, text.

Docling-backed formats live in tests/test_rag_parsing_docling.py so this file
runs in the API environment, where Docling is deliberately absent.
"""

import csv as _csv

import pytest
from openpyxl import Workbook

from app.rag.parsing import (
    ParseError,
    detect_file_type,
    parse_text_to_chunks,
    parse_to_chunks,
)

OPTS = {"max_chars": 500, "overlap_chars": 50}


@pytest.fixture()
def csv_file(tmp_path):
    p = tmp_path / "leave.csv"
    with p.open("w", newline="") as fh:
        w = _csv.writer(fh)
        w.writerow(["Employee", "Department", "Days"])
        for i in range(30):
            w.writerow([f"Person {i}", "HR", str(i)])
    return p


@pytest.fixture()
def xlsx_file(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Balances"
    ws.append(["Employee", "Department", "Days"])
    for i in range(30):
        ws.append([f"Person {i}", "HR", i])
    p = tmp_path / "balances.xlsx"
    wb.save(p)
    return p


@pytest.mark.parametrize("name,expected", [
    ("a.pdf", "pdf"), ("a.PDF", "pdf"), ("a.docx", "docx"),
    ("a.xlsx", "xlsx"), ("a.csv", "csv"), ("a.txt", "text"), ("a.md", "text"),
])
def test_detect_file_type(name, expected):
    assert detect_file_type(name) == expected


def test_unsupported_extension_is_rejected():
    with pytest.raises(ParseError):
        detect_file_type("malware.exe")


def test_text_parses_to_chunks():
    chunks = parse_text_to_chunks("A typed-in policy note.", **OPTS)
    assert len(chunks) == 1
    assert "typed-in policy" in chunks[0].content


def test_empty_typed_text_is_an_error_not_an_empty_document():
    """A document with zero chunks would be silently unsearchable."""
    with pytest.raises(ParseError):
        parse_text_to_chunks("   ", **OPTS)


def test_csv_chunks_repeat_the_header(csv_file):
    chunks = parse_to_chunks(csv_file, "csv", **OPTS)
    assert chunks
    assert all("Employee" in c.content for c in chunks)
    assert all(c.element_type == "table" for c in chunks)


def test_xlsx_chunks_cover_every_row(xlsx_file):
    """open_sheet_rows is uncapped, unlike load_table's ~200-row window."""
    chunks = parse_to_chunks(xlsx_file, "xlsx", **OPTS)
    joined = "\n".join(c.content for c in chunks)
    assert "Person 0" in joined
    assert "Person 29" in joined


def test_xlsx_covers_every_sheet(tmp_path):
    wb = Workbook()
    wb.active.title = "First"
    wb.active.append(["A"])
    wb.active.append(["one"])
    second = wb.create_sheet("Second")
    second.append(["B"])
    second.append(["two"])
    p = tmp_path / "multi.xlsx"
    wb.save(p)

    chunks = parse_to_chunks(p, "xlsx", **OPTS)
    joined = "\n".join(c.content for c in chunks)
    assert "First" in joined and "Second" in joined
    assert "one" in joined and "two" in joined


def test_chunk_indices_are_contiguous_across_sheets(tmp_path):
    wb = Workbook()
    wb.active.title = "S1"
    wb.active.append(["A"])
    for i in range(40):
        wb.active.append([f"r{i}"])
    s2 = wb.create_sheet("S2")
    s2.append(["B"])
    for i in range(40):
        s2.append([f"q{i}"])
    p = tmp_path / "two.xlsx"
    wb.save(p)

    chunks = parse_to_chunks(p, "xlsx", max_chars=120, overlap_chars=0)
    assert [c.chunk_index for c in chunks] == list(range(len(chunks)))


def test_a_corrupt_spreadsheet_raises_parse_error(tmp_path):
    p = tmp_path / "broken.xlsx"
    p.write_bytes(b"not a spreadsheet")
    with pytest.raises(ParseError):
        parse_to_chunks(p, "xlsx", **OPTS)


def test_an_empty_sheet_produces_no_chunks_and_raises(tmp_path):
    wb = Workbook()
    wb.active.append(["Header"])
    p = tmp_path / "headers_only.xlsx"
    wb.save(p)
    with pytest.raises(ParseError):
        parse_to_chunks(p, "xlsx", **OPTS)


def test_docling_is_not_imported_at_module_scope():
    """The API image must never pull torch. Docling IS installed in this venv,
    so this proves parsing.py avoids it by design rather than by absence.

    Runs in a SUBPROCESS deliberately: `sys.modules` is process-global, and
    tests/test_rag_parsing_docling.py imports docling in the same pytest run —
    an in-process check would pass or fail depending on test order rather than
    on the import graph, which is the thing actually under test.
    """
    import subprocess
    import sys

    probe = (
        "import sys; import app.rag.parsing;"
        "bad = [m for m in ('docling', 'torch') if m in sys.modules];"
        "print(','.join(bad))"
    )
    out = subprocess.run(
        [sys.executable, "-c", probe],
        capture_output=True, text=True, check=True,
    ).stdout.strip()
    assert out == "", f"importing app.rag.parsing pulled in: {out}"


def test_normalize_heading_folds_case_whitespace_and_trailing_punctuation():
    from app.rag.parsing import _normalize_heading

    assert _normalize_heading("  Table   of  Contents :  ") == "table of contents"
    assert _normalize_heading("CONTENTS.") == "contents"


def test_front_matter_is_skipped_by_first_segment():
    from app.rag.parsing import _is_skipped_section

    skip = {"table of contents", "contents", "index"}
    assert _is_skipped_section("Table of Contents", skip)
    assert _is_skipped_section("Table of Contents > 5.2.5 Assurance of Limits", skip)


def test_a_legitimate_index_section_is_not_skipped():
    """First-segment-only is the guard: a policy document's own
    'Index of Limits' under a real chapter must stay indexed."""
    from app.rag.parsing import _is_skipped_section

    skip = {"table of contents", "contents", "index"}
    assert not _is_skipped_section("Chapter 3 > Index of Limits", skip)
    assert not _is_skipped_section("Chapter 4: Investment Products", skip)


def test_skip_is_inert_with_no_section_or_empty_set():
    from app.rag.parsing import _is_skipped_section

    assert not _is_skipped_section(None, {"contents"})
    assert not _is_skipped_section("Contents", set())


def test_settings_expose_the_skip_list_normalized():
    from app.config import Settings

    s = Settings(rag_skip_sections="Table of Contents, Contents ,Index")
    assert s.rag_skipped_sections == {"table of contents", "contents", "index"}
    assert s.rag_chunk_min_body_chars == 40
