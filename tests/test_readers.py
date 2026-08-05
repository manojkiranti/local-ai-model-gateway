"""Offline tests for app/files/readers.py — the xlsx/csv normalizer.

Pure module: no DB, no HTTP. Builds tiny workbooks/CSVs on disk with openpyxl /
stdlib and asserts the normalized Table shape, multi-sheet inspection, cached
formula values (formulas are NEVER evaluated), and the row/char caps.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from app.files import readers


def _write_xlsx(path: Path, sheets: dict[str, list[list]]) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name)
        for row in rows:
            ws.append(row)
    wb.save(path)


# --------------------------------------------------------------------------- #
# inspect_workbook
# --------------------------------------------------------------------------- #
def test_inspect_reports_every_sheet(tmp_path):
    p = tmp_path / "book.xlsx"
    _write_xlsx(
        p,
        {
            "Q1": [["name", "amount"], ["a", 10], ["b", 20]],
            "Q2": [["name", "amount"], ["c", 30]],
            "Notes": [["free text"]],
        },
    )
    sheets = readers.inspect_workbook(p)
    assert [s.sheet_name for s in sheets] == ["Q1", "Q2", "Notes"]
    q1 = sheets[0]
    assert q1.headers == ["name", "amount"]
    assert q1.total_rows == 2        # data rows only (header excluded)
    assert q1.total_cols == 2
    assert len(q1.sample_rows) <= 10


def test_inspect_csv_is_one_pseudo_sheet(tmp_path):
    p = tmp_path / "people.csv"
    p.write_text("name,age\nAda,36\nGrace,45\n", encoding="utf-8")
    sheets = readers.inspect_workbook(p)
    assert len(sheets) == 1
    assert sheets[0].sheet_name == "people"
    assert sheets[0].headers == ["name", "age"]
    assert sheets[0].total_rows == 2  # data rows only (Ada, Grace)


# --------------------------------------------------------------------------- #
# load_table
# --------------------------------------------------------------------------- #
def test_load_table_default_first_sheet(tmp_path):
    p = tmp_path / "book.xlsx"
    _write_xlsx(p, {"Alpha": [["h"], ["1"]], "Beta": [["x"], ["2"]]})
    t = readers.load_table(p)
    assert t.sheet_name == "Alpha"
    assert t.headers == ["h"]


def test_load_table_select_sheet_by_name_caseinsensitive(tmp_path):
    p = tmp_path / "book.xlsx"
    _write_xlsx(p, {"Alpha": [["h"], ["1"]], "Beta": [["x"], ["2"]]})
    t = readers.load_table(p, sheet="beta")
    assert t.sheet_name == "Beta"
    assert t.rows[0] == ["2"]


def test_load_table_unknown_sheet_raises(tmp_path):
    p = tmp_path / "book.xlsx"
    _write_xlsx(p, {"Alpha": [["h"], ["1"]]})
    with pytest.raises(readers.SheetNotFound):
        readers.load_table(p, sheet="Nope")


def test_formulas_are_not_evaluated_cached_value_used(tmp_path):
    """data_only=True returns the last cached value; an unopened formula cell
    reads as None (never executed) — we must not crash or run it."""
    p = tmp_path / "calc.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["n"])
    ws["A2"] = "=1+1"  # never opened in Excel -> no cached value
    wb.save(p)
    t = readers.load_table(p)
    # The formula cell has no cached value -> empty; it is NOT "=1+1" and NOT "2".
    # (The lone all-empty data row is trimmed, so rows is empty — either way, no
    # formula text and no evaluated result ever appears.)
    flat = [c for row in t.rows for c in row]
    assert "=1+1" not in flat
    assert "2" not in flat


def test_row_cap_truncates_and_flags(tmp_path):
    p = tmp_path / "big.xlsx"
    rows = [["v"]] + [[i] for i in range(readers.READ_MAX_ROWS + 50)]
    _write_xlsx(p, {"S": rows})
    t = readers.load_table(p)
    assert t.truncated is True
    assert len(t.rows) <= readers.READ_MAX_ROWS


def test_start_row_paging(tmp_path):
    p = tmp_path / "page.xlsx"
    rows = [["v"]] + [[i] for i in range(10)]
    _write_xlsx(p, {"S": rows})
    t = readers.load_table(p, start_row=5, max_rows=3)
    # start_row is 1-based into the DATA rows: data = [0..9], 5th data row = "4".
    assert t.rows == [["4"], ["5"], ["6"]]


def test_column_projection(tmp_path):
    p = tmp_path / "cols.xlsx"
    _write_xlsx(p, {"S": [["a", "b", "c"], [1, 2, 3], [4, 5, 6]]})
    t = readers.load_table(p, columns=["a", "c"])
    assert t.headers == ["a", "c"]
    assert t.rows == [["1", "3"], ["4", "6"]]


def test_csv_utf8_sig_and_values(tmp_path):
    p = tmp_path / "b.csv"
    p.write_bytes("﻿name,city\nJi-woo,Seoul\n".encode("utf-8"))
    t = readers.load_table(p)
    assert t.headers == ["name", "city"]
    assert t.rows == [["Ji-woo", "Seoul"]]


def test_corrupt_file_raises_readerror(tmp_path):
    p = tmp_path / "bad.xlsx"
    p.write_bytes(b"this is not a zip/xlsx at all")
    with pytest.raises(readers.ReadError):
        readers.inspect_workbook(p)
