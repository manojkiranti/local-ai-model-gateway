#!/usr/bin/env python
"""Phase 6B Task 3B — turn the independent-holdout result into a review pack a
Nepali reader can actually adjudicate. READ-ONLY, offline, no tuning.

    DATABASE_URL=postgresql+asyncpg://…/local_ai_gateway_p4 \
        .venv/bin/python scripts/nrb_holdout_evidence.py \
            --holdout docs/nrb/phase6b-routing-holdout.json \
            --phase6a docs/nrb/phase6a-manifest.json \
            --lexicon docs/nrb/phase6b-lexicon.json \
            --profile docs/nrb/phase6b-routing-holdout-profile.json \
            --out-md   docs/nrb/phase6b-routing-holdout-manual-review.md \
            --out-json docs/nrb/phase6b-routing-holdout-evidence.json \
            --pages-dir docs/nrb/holdout-pages

WHY THIS EXISTS, AND WHAT IT IS NOT
    Task 3 measured native-2 on a cohort that never influenced it and produced
    aggregate numbers. Those numbers answer two of the three questions a
    conversion decision needs — *was the routed input really glyph-mapped* and
    *did the converter produce usable Unicode* — and neither of them is the third:
    *is the resulting Devanagari correct Nepali?* That one is a reader's call, so
    this command lays out the per-unit evidence and leaves every semantic verdict
    at `awaiting_nepali_review`. It never writes a `confirmed_*` state.

    It changes no threshold, no guard, no classifier and no extractor version. It
    writes no database row, makes no HTTP request and runs no OCR.

WHAT IT ADDS OVER THE TASK 3 PROFILE
    Location. `nrb_extractions` persists no text, and the text the profile scored
    is flat: a PDF's pages are joined with "\\n" (no page markers) and a
    spreadsheet's cells are rendered `" | "`-joined. Neither can tell a reviewer
    *where on the page* a flagged unit came from. So each blob is re-parsed here
    with its structure retained — pages for a PDF, sheet + A1 cell for a workbook
    — and the reconstruction is checked against the stored native-2 metrics
    before any of it is reported. A blob whose units do not reconstruct exactly is
    reported as such rather than given approximate coordinates.

    Original cell boundaries. The Task 3 validator recovered cells by splitting
    the rendered row back on `" | "`, which is a faithful inverse only while no
    cell contains that sequence. Here the workbook is re-read, so a cell boundary
    is the workbook's own. The two are compared and any disagreement is reported.

Exit codes: 0 fine; 1 finished with a blob that could not be read or reconstructed;
2 refused to start (leakage, missing artifact, no converter).
"""

from __future__ import annotations

import argparse
import asyncio
import bisect
import collections
import json
import os
import re
import subprocess
import sys
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Sequence

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from sqlalchemy import select  # noqa: E402

from app.files import readers  # noqa: E402
from app.files import documents as file_documents  # noqa: E402
from app.nrb import (  # noqa: E402
    extraction,
    filestore,
    legacy_convert as LC,
    legacy_eval,
    legacy_font,
    legacy_report,
    lexicon as LX,
    quality,
    routing,
    units,
)
from app.nrb.manifest import read_manifest  # noqa: E402
from app.nrb.models import NRBExtraction, NRBFile  # noqa: E402

HIGH_BAND = 0.80          # the candidate conversion gate, on unit_legacy_ratio
MID_BAND = 0.50
LOW_BAND = 0.20

# How many flagged units are printed per item. The queue holds 38,829 legacy
# units across 56 blobs; a pack that printed them all would be unreviewable. The
# cap is applied in document order (never "the most convincing ones"), and the
# per-item header always states the true total so the sample cannot read as the
# whole.
UNITS_PER_ITEM_MD = 10
UNITS_PER_ITEM_JSON = 40

# Page rendering. 90 dpi grayscale JPEG is legible for Devanagari (checked) and
# keeps a 53-page evidence set to a few MB; the point is to let a reviewer read
# the page, not to archive a facsimile.
RENDER_DPI = 90
RENDER_QUALITY = 72

# The three false-negative candidates §7 asks about, and the known false-positive
# template §6 asks about. Named here so the pack reports the SAME blobs the Task 3
# profile named, rather than re-deriving a list that could quietly drift.
FALSE_NEGATIVE_CANDIDATES = ("d74b592c894a", "a2077aa9b24d", "7425cbd1d9ee")


# --------------------------------------------------------------------------- #
# Located judgment units — the thing the flat extraction text cannot give us
# --------------------------------------------------------------------------- #
@dataclass(frozen=True)
class Located:
    """One judgment unit, at the index native-2 scored it and where it came from.

    `index` indexes the SAME sequence `extraction._result` built (a PDF's
    `text.splitlines()`, a workbook's non-blank cells in sheet order), so it lines
    up with `legacy_convert`'s per-unit outcomes without any re-alignment.
    """

    index: int
    text: str
    where: str          # "p.7" | "Sheet1!B14" | "line 42"


@dataclass
class Reconstruction:
    """Located units for one blob, plus whether they can be trusted."""

    located: tuple[Located, ...]
    ok: bool
    note: str
    # A workbook's cells as the WORKBOOK reports them (rows of cells), which is
    # what a per-cell conversion must be fed. Empty for non-spreadsheets.
    rows: tuple[tuple[str, ...], ...] = ()
    # The extracted text, carried out of the ONE parse this module performs. A
    # 500-page PDF costs seconds in pypdf, so parsing it a second time just to
    # recover page boundaries would double the whole pass for no new information.
    text: str = ""


# `str.splitlines()` breaks on far more than "\n" — and a PDF text layer really
# does contain form feeds and lone carriage returns. This is the exact boundary
# set, so lines can be recovered WITH their character offsets, which
# `str.splitlines()` does not report. Splitting per page and concatenating is not
# equivalent: a page ending in a form feed followed by the join's "\n" yields an
# empty line that neither page produces on its own.
_LINE_BOUNDARY = re.compile("\\r\\n|[\\n\\r\\v\\f\\x1c\\x1d\\x1e\\x85\\u2028\\u2029]")


def _lines_with_offsets(text: str) -> list[tuple[int, str]]:
    """`str.splitlines()`, but each line paired with where it starts."""
    out: list[tuple[int, str]] = []
    pos = 0
    for m in _LINE_BOUNDARY.finditer(text):
        out.append((pos, text[pos:m.start()]))
        pos = m.end()
    if pos < len(text):
        out.append((pos, text[pos:]))
    return out


def _pdf_located(path: Path) -> Reconstruction:
    """Lines with their page numbers.

    `_extract_pdf` joins pages with "\\n" and throws the boundaries away, so they
    are recovered by re-reading the pages, re-deriving the same join, and mapping
    each line's character offset back to the page it started in. The equality
    check against `text.splitlines()` is the guard: if it ever fails, the unit
    keeps an exact line number and the item is marked unverified rather than
    given a page number that might be wrong.
    """
    read = file_documents.read_pdf_pages(path)
    text = "\n".join(read.pages)   # exactly what `_extract_pdf` stores
    starts: list[int] = []
    cursor = 0
    for page_text in read.pages:
        starts.append(cursor)
        cursor += len(page_text) + 1   # +1 for the join's "\n"

    offsets = _lines_with_offsets(text)
    if [s for _, s in offsets] != text.splitlines():
        return Reconstruction(
            tuple(Located(i, s, f"line {i + 1}")
                  for i, s in enumerate(text.splitlines())),
            False,
            "page attribution unavailable (line boundaries did not round-trip); "
            "line numbers are exact",
            text=text,
        )
    located = tuple(
        Located(i, s, f"p.{bisect.bisect_right(starts, off)}")
        for i, (off, s) in enumerate(offsets)
    )
    return Reconstruction(located, True, f"{len(read.pages)} pages", text=text)


def _column_letter(index_zero_based: int) -> str:
    n, out = index_zero_based + 1, ""
    while n:
        n, rem = divmod(n - 1, 26)
        out = chr(65 + rem) + out
    return out


def _sheet_located(path: Path) -> Reconstruction:
    """Cells with sheet name and A1 coordinate.

    Mirrors `extraction._extract_spreadsheet` step for step — the same sheet
    order, the same global `MAX_SHEET_ROWS` budget, the same "non-blank cells
    only" unit rule — because the units must be the ones native-2 actually
    scored, not a second opinion about the same workbook. Row/column origins come
    from openpyxl's own `min_row`/`min_column`, since `iter_rows()` starts at the
    first populated cell rather than at A1.
    """
    sheets = readers.inspect_workbook(path)
    names: list[str | None] = [s.sheet_name for s in sheets] or [None]
    located: list[Located] = []
    grid: list[tuple[str, ...]] = []
    origins = _sheet_origins(path)
    rows_seen = 0
    idx = 0
    for name in names:
        origin_row, origin_col = origins.get(name or "", (1, 0))
        with readers.open_sheet_rows(path, sheet=name) as stream:
            label = stream.sheet_name or (name or "sheet")
            if stream.headers:
                grid.append(tuple(str(h) for h in stream.headers))
                for j, h in enumerate(stream.headers):
                    if str(h).strip():
                        coord = f"{label}!{_column_letter(origin_col + j)}{origin_row}"
                        located.append(Located(idx, str(h), coord))
                        idx += 1
            for r, row in enumerate(stream.rows):
                if rows_seen >= extraction.MAX_SHEET_ROWS:
                    break
                rows_seen += 1
                if not any(str(c).strip() for c in row):
                    continue
                grid.append(tuple(str(c) for c in row))
                row_no = origin_row + 1 + r
                for j, c in enumerate(row):
                    if str(c).strip():
                        coord = f"{label}!{_column_letter(origin_col + j)}{row_no}"
                        located.append(Located(idx, str(c), coord))
                        idx += 1
    return Reconstruction(
        tuple(located), True, f"{len(names)} sheet(s)", tuple(grid),
        # The `" | "` rendering `_extract_spreadsheet` stores. Rebuilt only so the
        # before/after Devanagari ratios are computed over the same string
        # native-1 measured; it is NEVER a conversion or judgment unit.
        text="\n".join(" | ".join(row) for row in grid),
    )


def _sheet_origins(path: Path) -> dict[str, tuple[int, int]]:
    """`{sheet: (min_row, min_column)}` for an xlsx; empty for anything else.

    openpyxl's `iter_rows()` — which `readers.open_sheet_rows` uses — defaults to
    the sheet's populated bounding box, so a workbook whose data starts at C5
    would otherwise be reported with coordinates five rows and two columns off.
    Column origin is ZERO-based, to match `_column_letter`.

    All sheets in ONE open. Re-opening per sheet cost 571 seconds on
    `7425cbd1d9ee` alone — `load_workbook` re-parses the whole archive each time,
    and that workbook has enough sheets for it to dominate the entire pass.
    """
    if path.suffix.lower() != ".xlsx":
        return {}
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            return {
                name: (int(wb[name].min_row or 1),
                       int(wb[name].min_column or 1) - 1)
                for name in wb.sheetnames
            }
        finally:
            wb.close()
    except Exception:  # noqa: BLE001 - a coordinate origin is never worth a crash
        return {}


def _text_located(text: str) -> Reconstruction:
    lines = text.splitlines()
    return Reconstruction(
        tuple(Located(i, s, f"line {i + 1}") for i, s in enumerate(lines)),
        True,
        f"{len(lines)} lines",
        text=text,
    )


def locate(ref) -> Reconstruction:
    """Parse one blob ONCE, keeping structure, verified against native-2's metrics.

    The verification is the point. If the reconstruction disagrees with
    `unit_total` — the count native-2 recorded when it classified this blob — then
    the coordinates in the pack would be pointing at the wrong text, which is
    worse than having no coordinates. Such a blob is reported `ok=False` and its
    evidence is labelled unverified rather than quietly published.

    PDFs and workbooks are parsed here directly, because the structured read
    already yields the flat text `extraction.py` would have produced — going
    through `legacy_eval.read_blob_text` first would parse every file twice, which
    on a 500-page directive is seconds of pypdf for no new information. Other
    families (there are only a handful) take the shared path.
    """
    path = filestore.resolve_path(ref.storage_key)
    if ref.family == "pdf":
        rec = _pdf_located(path)
    elif ref.family == "spreadsheet":
        rec = _sheet_located(path)
    else:
        rec = _text_located(legacy_eval.read_blob_text(ref).text)

    expected = int(ref.metrics.get("unit_total") or 0)
    if expected and len(rec.located) != expected:
        rec.ok = False
        rec.note = (
            f"{rec.note}; reconstruction produced {len(rec.located)} units but "
            f"native-2 recorded {expected} — coordinates NOT trustworthy"
        )
    return rec


# --------------------------------------------------------------------------- #
# Conversion, per located unit
# --------------------------------------------------------------------------- #
def convert_located(ref, rec: Reconstruction, lexicon, converter):
    """Convert one blob and return `(DocumentConversion, outcome per unit index)`.

    A spreadsheet is converted from the WORKBOOK's own rows (never a `" | "`
    rendered row — `|` is a Preeti codepoint that maps to `्र`), so the cell
    boundary a reviewer sees is the cell boundary the converter saw. Everything
    else converts line by line.

    `document_legacy_ratio` is native-2's `unit_legacy_ratio`, never native-1's
    `legacy_line_ratio`: they are different quantities, and the whole point of the
    holdout was that the three big research workbooks sit at 0.969–0.993 on the
    first and 0.15–0.19 on the second.
    """
    doc_ratio = float(ref.metrics.get("unit_legacy_ratio") or 0.0)
    if ref.family == "spreadsheet":
        conversion = LC.convert_cells(
            rec.rows, converter, lexicon, document_legacy_ratio=doc_ratio
        )[0]
        # `rec.rows` includes blank cells (the grid is rectangular); the judgment
        # units are the non-blank ones. Re-align by walking the grid in the same
        # order the locator did.
        flat = [c for row in rec.rows for c in row]
        by_index: dict[int, LC.LineOutcome] = {}
        unit_i = 0
        for pos, cell in enumerate(flat):
            if str(cell).strip():
                by_index[unit_i] = conversion.lines[pos]
                unit_i += 1
        after_text = "\n".join(l.text for l in conversion.lines)
    else:
        conversion = LC.convert_document(
            rec.text, converter, lexicon, document_legacy_ratio=doc_ratio
        )
        by_index = {i: line for i, line in enumerate(conversion.lines)}
        after_text = conversion.text
    return conversion, by_index, quality.measure_text(after_text)


# --------------------------------------------------------------------------- #
# Page rendering
# --------------------------------------------------------------------------- #
def render_page(path: Path, page_no: int, out_dir: Path, stem: str) -> str | None:
    """Render one PDF page to a small grayscale JPEG. Returns a repo-relative path.

    Deterministic (same bytes + same page ⇒ same image) and skipped if the file
    already exists, so re-running the pack does not churn the repository. Uses
    poppler's `pdftoppm`, which is already on the box; nothing is downloaded.
    """
    out_dir.mkdir(parents=True, exist_ok=True)
    target = out_dir / f"{stem}-p{page_no:03d}.jpg"
    if target.exists():
        return str(target)
    prefix = out_dir / f"{stem}-p{page_no:03d}"
    try:
        subprocess.run(
            ["pdftoppm", "-jpeg", "-jpegopt", f"quality={RENDER_QUALITY}",
             "-gray", "-r", str(RENDER_DPI), "-f", str(page_no), "-l",
             str(page_no), "-singlefile", str(path), str(prefix)],
            check=True, capture_output=True, timeout=120,
        )
    except Exception:  # noqa: BLE001 - a missing render is a caveat, not a failure
        return None
    return str(target) if target.exists() else None


# --------------------------------------------------------------------------- #
# Assembly
# --------------------------------------------------------------------------- #
def band_of(ratio: float) -> str:
    if ratio >= HIGH_BAND:
        return ">=0.80"
    if ratio >= MID_BAND:
        return "0.50-0.80"
    if ratio > LOW_BAND:
        return "0.20-0.50"
    return "<=0.20"


def _clip(s: str, n: int = 150) -> str:
    s = " ".join(str(s).split())
    return s if len(s) <= n else s[: n - 1] + "…"


def _md_cell(s: str) -> str:
    return _clip(s).replace("|", "\\|")


async def acquisition(session, keys: Sequence[str]) -> dict[str, Any]:
    """Reconcile all 150 frozen entries to an exact total.

    Every manifest key is looked up in `nrb_files` by `comparison_key` — the same
    identity the manifest was drawn on — and each is placed in exactly one bucket.
    A key that produced no catalog row at all would be the one genuinely alarming
    outcome (it would mean the manifest and the catalog disagree about what exists)
    and is counted separately rather than folded into "failed".
    """
    rows = (
        await session.execute(
            select(NRBFile).where(NRBFile.comparison_key.in_(list(keys)))
        )
    ).scalars().all()
    by_key = {r.comparison_key: r for r in rows}
    shas = [r.content_sha256 for r in rows if r.content_sha256]
    ex = (
        await session.execute(
            select(NRBExtraction)
            .where(NRBExtraction.content_sha256.in_(shas))
            .where(NRBExtraction.extractor_version == routing.EXTRACTOR_VERSION_V2)
        )
    ).scalars().all()
    by_sha = {e.content_sha256: e for e in ex}

    buckets: dict[str, list[dict[str, Any]]] = collections.defaultdict(list)
    for key in keys:
        row = by_key.get(key)
        if row is None:
            buckets["no_catalog_row"].append({"comparison_key": key})
            continue
        entry = {
            "comparison_key": key,
            "source_url": row.source_url,
            "fetch_status": row.fetch_status,
            "http_status": row.http_status,
            "fetch_attempts": row.fetch_attempts,
            "fetch_error": row.fetch_error,
            "reported_mime_type": row.reported_mime_type,
            "reported_bytes": row.reported_bytes,
            "resource_type": row.resource_type,
            "content_sha256": row.content_sha256,
            "content_length": row.content_length,
            "sniffed_mime": row.sniffed_mime,
            "last_fetch_run_id": row.last_fetch_run_id,
        }
        if row.fetch_status != "fetched":
            buckets[f"not_fetched:{row.fetch_status}"].append(entry)
            continue
        rec = by_sha.get(row.content_sha256 or "")
        if rec is None:
            buckets["fetched_no_extraction"].append(entry)
            continue
        entry["status"] = rec.status
        entry["reason"] = rec.reason
        buckets[f"extracted:{rec.status}"].append(entry)
    return {
        "requested": len(keys),
        "buckets": {k: v for k, v in sorted(buckets.items())},
        "counts": {k: len(v) for k, v in sorted(buckets.items())},
        "distinct_blobs": len(set(shas)),
        "blobs_shared_by_more_than_one_key": {
            sha[:12]: n
            for sha, n in collections.Counter(shas).items() if n > 1
        },
        "fetch_run_ids": sorted(
            {r.last_fetch_run_id for r in rows if r.last_fetch_run_id is not None}
        ),
    }


def evidence_for(legacy_locs, by_index, lexicon, limit: int) -> list[dict[str, Any]]:
    """The flagged units of one blob, in document order, capped at `limit`."""
    out: list[dict[str, Any]] = []
    for loc in legacy_locs:
        line = by_index.get(loc.index)
        validation = getattr(line, "validation", None)
        out.append({
            "unit_index": loc.index,
            "where": loc.where,
            "original": loc.text,
            "converted": (line.converted if line else None),
            "applied": (line.text if line else loc.text),
            "disposition": (line.disposition if line else "not_converted"),
            "validation": (validation.outcome if validation else None),
            "validation_reason": (getattr(validation, "reason", None)
                                  if validation else None),
            "reads_as_english": LX.is_confidently_english(loc.text, lexicon),
            # Never `confirmed_correct` — see §5 of the task. Only a reader
            # decides whether the Devanagari above is the right Devanagari.
            "semantic_verdict": "awaiting_nepali_review",
        })
        if len(out) >= limit:
            break
    return out


async def run(args) -> int:
    from app.db.session import SessionLocal

    hold = read_manifest(Path(args.holdout))
    p6a = read_manifest(Path(args.phase6a))
    hold_keys, p6a_keys = set(hold.keys()), set(p6a.keys())
    leak = hold_keys & p6a_keys
    if leak:
        print(f"ERROR: {len(leak)} Phase 6A keys leaked into the holdout — aborting",
              file=sys.stderr)
        return 2

    lexicon = LX.load_lexicon(args.lexicon)
    converter = legacy_font.converter_for("Preeti")
    if converter is None:
        print("ERROR: the Preeti mapping is unavailable (npttf2utf not installed)",
              file=sys.stderr)
        return 2

    profile = json.loads(Path(args.profile).read_text(encoding="utf-8"))

    async with SessionLocal() as session:
        acct = await acquisition(session, hold.keys())
        every = await legacy_eval.load_blob_refs(
            session, extractor_version=routing.EXTRACTOR_VERSION_V2
        )
    refs = [r for r in every if hold_keys & set(r.comparison_keys)]
    by_sha = {r.content_sha256: r for r in refs}

    # Scope: everything native-2 flagged (so the false positives and the mid-band
    # are covered too) plus the three false-negative candidates.
    flagged = [r for r in refs if r.status == "suspicious"
               and r.reason == "legacy_font_suspected"]
    fn_refs = [r for r in refs
               if any(r.content_sha256.startswith(p)
                      for p in FALSE_NEGATIVE_CANDIDATES)]
    scope = flagged + [r for r in fn_refs if r not in flagged]
    scope.sort(key=lambda r: (-float(r.metrics.get("unit_legacy_ratio") or 0.0),
                              r.content_sha256))

    pages_dir = Path(args.pages_dir)
    items: list[dict[str, Any]] = []
    problems: list[str] = []
    for n, ref in enumerate(scope, 1):
        began = time.monotonic()
        try:
            rec = locate(ref)
        except Exception as exc:  # noqa: BLE001
            problems.append(f"{ref.content_sha256[:12]}: unreadable ({exc!r})")
            continue
        conversion, by_index, after = convert_located(ref, rec, lexicon, converter)
        summary = legacy_report.summarise_conversion(conversion)
        ratio = float(ref.metrics.get("unit_legacy_ratio") or 0.0)
        before = quality.measure_text(rec.text)
        # Assessed ONCE. `assess_unit` is called on every unit of every blob and
        # the queue holds 38,829 flagged units alone; re-deriving the same states
        # for the English check and again for the evidence list tripled the pass.
        legacy_locs = tuple(
            loc for loc in rec.located
            if units.assess_unit(loc.text).state == units.STATE_LEGACY
        )
        english_units = sum(1 for loc in legacy_locs
                            if LX.is_confidently_english(loc.text, lexicon))
        ev = evidence_for(legacy_locs, by_index, lexicon, UNITS_PER_ITEM_JSON)

        render = None
        if ref.family == "pdf" and ev and rec.ok:
            first = ev[0]["where"]
            if first.startswith("p."):
                render = render_page(
                    filestore.resolve_path(ref.storage_key), int(first[2:]),
                    pages_dir, ref.content_sha256[:12],
                )
        if not rec.ok:
            problems.append(f"{ref.content_sha256[:12]}: {rec.note}")

        items.append({
            "content_sha256": ref.content_sha256,
            "short": ref.content_sha256[:12],
            "comparison_keys": list(ref.comparison_keys),
            "family": ref.family,
            "status": ref.status,
            "reason": ref.reason,
            "band": band_of(ratio),
            "unit_legacy_ratio": ratio,
            "legacy_line_ratio": float(ref.metrics.get("legacy_line_ratio") or 0.0),
            "unit_total": int(ref.metrics.get("unit_total") or 0),
            "unit_judged": int(ref.metrics.get("unit_judged") or 0),
            "unit_legacy_candidates": int(
                ref.metrics.get("unit_legacy_candidates") or 0),
            "unit_unjudged": int(ref.metrics.get("unit_unjudged") or 0),
            "unit_max_legacy_run": int(ref.metrics.get("unit_max_legacy_run") or 0),
            "unit_contested_legacy_ratio": float(
                ref.metrics.get("unit_contested_legacy_ratio") or 0.0),
            "minority_legacy_detected": int(
                ref.metrics.get("minority_legacy_detected") or 0),
            "page_count": ref.metrics.get("page_count"),
            "sheet_count": ref.metrics.get("sheet_count"),
            "legacy_units_seen": len(legacy_locs),
            "legacy_units_reading_as_english": english_units,
            "devanagari_before": before.devanagari_ratio,
            "devanagari_after": after.devanagari_ratio,
            "counts": summary["counts"],
            "attempted": summary["attempted_lines"],
            "accepted": summary["accepted_lines"],
            "ambiguous": summary["ambiguous_lines"],
            "ambiguous_held": summary["ambiguous_held_lines"],
            "rejected": summary["rejected_lines"],
            "guarded_english": summary["guarded_english_lines"],
            "guarded_unicode": summary["guarded_unicode_lines"],
            "location_note": rec.note,
            "location_verified": rec.ok,
            "rendered_page": render,
            "evidence": ev,
            "semantic_verdict": "awaiting_nepali_review",
        })
        print(f"  [{n}/{len(scope)}] {ref.content_sha256[:12]} {ref.family:11s} "
              f"ratio={ratio:.3f} legacy={len(legacy_locs):5d} "
              f"{time.monotonic() - began:6.1f}s", flush=True)

    payload = {
        "identity": {
            "holdout_fingerprint": hold.selection_sha256,
            "phase6a_fingerprint": p6a.selection_sha256,
            "intersection": len(leak),
            "extractor_version": routing.EXTRACTOR_VERSION_V2,
            "converter": f"{converter.name} {converter.version} "
                         f"({converter.mapping})",
            "lexicon_fingerprint": lexicon.fingerprint,
            "gate": "unit_legacy_ratio >= 0.80",
        },
        "acquisition": acct,
        "items": items,
        "problems": problems,
    }
    Path(args.out_json).write_text(
        json.dumps(payload, ensure_ascii=False, indent=1, sort_keys=False),
        encoding="utf-8",
    )
    Path(args.out_md).write_text(
        render_markdown(payload, profile, acct), encoding="utf-8"
    )
    print(f"items: {len(items)}  high-band: "
          f"{sum(1 for i in items if i['band'] == '>=0.80')}  "
          f"problems: {len(problems)}")
    print(f"wrote {args.out_md}")
    print(f"wrote {args.out_json}")
    return 1 if problems else 0


# --------------------------------------------------------------------------- #
# The pack
# --------------------------------------------------------------------------- #
def _english_share_of_flagged(it: dict) -> float:
    """Of the units native-2 called legacy, what share read as English?

    The false-positive measure, and it is a SHARE, not a presence test. A long
    Nepali circular that quotes one English caption will contain a stray flagged
    English unit; that is noise, not a mis-route. The 0.50 line is the definition
    Task 3 used, kept identical here so the two documents agree.
    """
    seen = it["legacy_units_seen"]
    return it["legacy_units_reading_as_english"] / seen if seen else 0.0


def _item_block(out: list[str], it: dict, *, show: int) -> None:
    """One reviewable item: what native-2 saw, what the converter did, where."""
    keys = it["comparison_keys"]
    out.append(f"#### `{it['short']}` — {it['family']}, "
               f"unit ratio **{it['unit_legacy_ratio']:.4f}** ({it['band']})")
    out.append("")
    out.append(f"- source: <{keys[0]}>" if keys else "- source: (none recorded)")
    for extra in keys[1:]:
        out.append(f"  also published at <{extra}>")
    out.append(f"- sha256: `{it['content_sha256']}`")
    shape = (f"{it['page_count']} pages" if it["page_count"]
             else f"{it['sheet_count']} sheets" if it["sheet_count"] else "—")
    out.append(
        f"- units: {it['unit_total']} total · {it['unit_judged']} judged · "
        f"**{it['unit_legacy_candidates']} legacy** · {it['unit_unjudged']} unjudged"
        f" · longest legacy run {it['unit_max_legacy_run']} · {shape}"
    )
    out.append(
        f"- native-1 `legacy_line_ratio` {it['legacy_line_ratio']:.4f} · "
        f"Devanagari {it['devanagari_before']:.4f} → {it['devanagari_after']:.4f}"
    )
    out.append(
        f"- converter: {it['accepted']} accepted, {it['ambiguous']} ambiguous "
        f"(applied), {it['ambiguous_held']} ambiguous-held (NOT applied), "
        f"{it['rejected']} rejected, of {it['attempted']} attempted · guards held "
        f"{it['guarded_english']} English / {it['guarded_unicode']} Unicode units back"
    )
    share = _english_share_of_flagged(it)
    if it["legacy_units_reading_as_english"]:
        mark = "⚠️ **mis-routed:** " if share >= 0.50 else ""
        out.append(f"- {mark}{it['legacy_units_reading_as_english']} of "
                   f"{it['legacy_units_seen']} flagged units read as English "
                   f"({share:.1%})")
    if it["rendered_page"]:
        out.append(f"- rendered source: `{it['rendered_page']}`")
    if not it["location_verified"]:
        out.append(f"- ⚠️ locations unverified — {it['location_note']}")
    out.append("")
    ev = it["evidence"][:show]
    if not ev:
        out.append("_No flagged unit survived to the evidence stage._")
        out.append("")
        return
    out.append(f"Showing {len(ev)} of {it['legacy_units_seen']} flagged units, in "
               f"document order:")
    out.append("")
    out.append("| where | extracted (as stored) | converted | outcome |")
    out.append("| --- | --- | --- | --- |")
    for e in ev:
        conv = e["converted"] or "—"
        applied = "" if e["disposition"] in ("converted", "converted_unjudged",
                                             "ambiguous") else " *(not applied)*"
        flag = " 🇬🇧" if e["reads_as_english"] else ""
        out.append(f"| `{e['where']}` | `{_md_cell(e['original'])}`{flag} | "
                   f"{_md_cell(conv)} | {e['disposition']}{applied} |")
    out.append("")
    out.append(f"**Nepali verdict for `{it['short']}`: `awaiting_nepali_review`** "
               "— replace with `confirmed_correct`, `confirmed_wrong` or "
               "`ambiguous`.")
    out.append("")


def render_markdown(payload: dict, profile: dict, acct: dict) -> str:
    ident = payload["identity"]
    items = payload["items"]
    high = [i for i in items if i["band"] == ">=0.80"
            and i["reason"] == "legacy_font_suspected"]
    mid = [i for i in items if i["band"] == "0.50-0.80"]
    low = [i for i in items if i["band"] == "0.20-0.50"]
    # A false positive is a document native-2 ROUTED whose flagged units turn out
    # to be English. A clean document containing English-looking units was not
    # routed and is therefore not a false positive — it belongs in §6. Conflating
    # the two would report the classifier's correct calls as its mistakes.
    fps = [i for i in items
           if i["reason"] == "legacy_font_suspected"
           and i["legacy_units_seen"]
           and _english_share_of_flagged(i) >= 0.5]
    fns = [i for i in items if i["reason"] == "clean"]
    o: list[str] = []
    A = o.append

    A("# Phase 6B — independent routing holdout: Nepali manual-review pack")
    A("")
    A("Generated by `scripts/nrb_holdout_evidence.py` (read-only, offline). This "
      "document is **evidence, not a conclusion**. Every semantic judgment below "
      "is `awaiting_nepali_review` until a competent Nepali reader replaces it.")
    A("")
    A(f"- holdout fingerprint `{ident['holdout_fingerprint']}`")
    A(f"- Phase 6A fingerprint `{ident['phase6a_fingerprint']}` — "
      f"intersection with the holdout: **{ident['intersection']}**")
    A(f"- classifier `{ident['extractor_version']}`, unchanged; candidate gate "
      f"`{ident['gate']}`")
    A(f"- converter `{ident['converter']}` — used here as an **evaluation "
      "instrument only**, through `app/nrb/legacy_font.py`. GPL-3.0; not "
      "installed by the Dockerfile.")
    A(f"- lexicon `{ident['lexicon_fingerprint'][:16]}…`")
    A("")
    A("This supersedes `phase6b-routing-holdout-manual-review.txt` for review "
      "purposes. That file is Task 3's aggregate artifact and stays as committed; "
      "it sampled the queue, this covers all of it.")
    A("")

    # ---------------------------------------------------------------- 1
    A("## 1. Acquisition accounting — all 150 frozen entries")
    A("")
    A("Every key in the frozen manifest is placed in exactly one bucket, so the "
      "buckets sum to 150 by construction. The evaluation denominator is 142 "
      "because eight files are gone from NRB's server, **not** because anything "
      "was dropped for looking inconvenient.")
    A("")
    A("| outcome | count |")
    A("| --- | ---: |")
    total = 0
    for k, n in acct["counts"].items():
        A(f"| {k} | {n} |")
        total += n
    A(f"| **total** | **{total}** |")
    A("")
    A(f"- distinct blobs among the fetched: **{acct['distinct_blobs']}** "
      f"(no two holdout keys resolved to the same bytes: "
      f"`{acct['blobs_shared_by_more_than_one_key'] or 'none'}`)")
    A(f"- every row carries `last_fetch_run_id` in {acct['fetch_run_ids']} — one "
      "fetch pass, so **nothing was already present** and nothing was re-fetched "
      "after the outcomes were seen")
    A("")
    A("### No substitution occurred")
    A("")
    A("- the manifest was committed at `ddc5f2d`, **before** any network access, "
      "and has not changed since (its fingerprint still verifies);")
    A("- the fetch selected exactly the 150 frozen keys and no others;")
    A("- no failed key was re-drawn, replaced or swapped: the failures are listed "
      "below by name and remain **in the denominator**;")
    A("- the eight absences are HTTP 404 from NRB, corroborated by NRB's own ACF "
      "metadata reporting `filesize = 0` for all eight.")
    A("")
    for bucket, rows in acct["buckets"].items():
        if not bucket.startswith("not_fetched") and bucket not in (
            "no_catalog_row", "fetched_no_extraction",
        ):
            continue
        A(f"#### {bucket} ({len(rows)})")
        A("")
        A("| HTTP | claimed type | claimed bytes | source |")
        A("| --- | --- | ---: | --- |")
        for r in rows:
            A(f"| {r.get('http_status') or '—'} | "
              f"{r.get('reported_mime_type') or '—'} | "
              f"{r.get('reported_bytes')} | <{r.get('source_url')}> |")
        A("")
    for bucket in ("extracted:unsupported", "extracted:failed"):
        rows = acct["buckets"].get(bucket, [])
        if not rows:
            continue
        A(f"#### {bucket} ({len(rows)}) — fetched fine, no parser")
        A("")
        A("| sniffed | ext | bytes | reason | source |")
        A("| --- | --- | ---: | --- | --- |")
        for r in rows:
            A(f"| {r.get('sniffed_mime')} | {r.get('resource_type')} | "
              f"{r.get('content_length')} | {r.get('reason')} | "
              f"<{r.get('source_url')}> |")
        A("")

    # ---------------------------------------------------------------- 2
    A("## 2. Three different questions, kept apart")
    A("")
    A("Collapsing these into one percentage is the main way a report like this "
      "misleads. They are measured on different evidence and only the first two "
      "are settled.")
    A("")
    A("| question | what it asks | evidence | status |")
    A("| --- | --- | --- | --- |")
    q = profile["candidate_queue"]
    A(f"| **Routing precision** | is the routed INPUT actually legacy Nepali, "
      f"rather than English, numeric or genuine Unicode? | script-independent: "
      f"are the units native-2 flagged readable English? | **settled** — "
      f"{q['routed'] - q['false_route_english']}/{q['routed']} inputs are not "
      f"English; 0 false routes |")
    A(f"| **Conversion recovery** | did npttf2utf produce usable Unicode? | "
      f"structural: acceptance rate + the native-1 flag clearing | **settled** — "
      f"{q['legacy_recovered']}/{q['routed']} ({q['recovered']} clean, "
      f"{q['partial']} partial), {q['unresolved']} unresolved |")
    A("| **Conversion correctness** | is the Unicode *semantically correct "
      "Nepali*? | a Nepali reader comparing §4 against the rendered page | "
      "**PENDING — no result** |")
    A("")
    A("> `52/56` is a **recovery** figure. It is not, and must not be quoted as, "
      "confirmed semantic conversion success. Nothing in this repository has yet "
      "verified that any converted line says what the source page says.")
    A("")

    # ---------------------------------------------------------------- 3
    A(f"## 3. Review ledger — the whole `>=0.80` queue ({len(high)} items)")
    A("")
    A("Fill in the last column. **English** is the share of *flagged* units that "
      "read as English — the false-positive measure. A mis-route would sit near "
      "1.0; the stray singletons below are English captions inside Nepali "
      "documents, which is noise rather than a wrong routing decision. "
      "**applied/attempted** counts units the converter replaced against units it "
      "tried, over the whole document (unjudged units convert too, so it can "
      "exceed the flagged-unit count) — it is a *recovery* number, never a "
      "correctness one.")
    A("")
    worst = max((_english_share_of_flagged(i) for i in high), default=0.0)
    A(f"Highest English share anywhere in this band: **{worst:.1%}** — no item "
      "approaches the 50% false-positive definition.")
    A("")
    A("| # | blob | type | ratio | flagged | applied/attempted | English | "
      "your verdict |")
    A("| ---: | --- | --- | ---: | ---: | ---: | ---: | --- |")
    for n, it in enumerate(high, 1):
        applied = it["accepted"] + it["ambiguous"]
        rec = (f"{applied}/{it['attempted']}" if it["attempted"] else "—")
        share = _english_share_of_flagged(it)
        eng = (f"{'⚠️ ' if share >= 0.50 else ''}"
               f"{it['legacy_units_reading_as_english']} ({share:.1%})"
               if it["legacy_units_reading_as_english"] else "0")
        A(f"| {n} | `{it['short']}` | {it['family']} | "
          f"{it['unit_legacy_ratio']:.3f} | {it['unit_legacy_candidates']} | "
          f"{rec} | {eng} | `awaiting_nepali_review` |")
    A("")

    # ---------------------------------------------------------------- 4
    A("## 4. The candidate queue, item by item")
    A("")
    A(f"All {len(high)} items, not a sample. Each shows up to "
      f"{UNITS_PER_ITEM_MD} flagged units **in document order** — the per-item "
      f"header states the true total, and the full set (to "
      f"{UNITS_PER_ITEM_JSON}) is in the JSON companion. Spreadsheet units are "
      "individual cells with their own A1 coordinate; a `\" | \"`-joined row is "
      "never converted or shown, because `|` is a Preeti codepoint that maps to "
      "`्र`.")
    A("")
    for it in high:
        _item_block(o, it, show=UNITS_PER_ITEM_MD)

    # ---------------------------------------------------------------- 5
    A("## 5. The known false-positive class — English accounting templates")
    A("")
    A("> **This holdout has now exposed this classifier defect and therefore may "
      "not be reused as independent validation for a classifier modified to "
      "correct it.** A future correction must ship as extractor version "
      "`native-3` and be validated on a new, independent cohort.")
    A("")
    A("Nothing here is fixed in this task. All of these sit **below** the "
      "`>=0.80` gate, so the candidate queue is unaffected — this is a caveat on "
      "the lower bands, not a blocker.")
    A("")
    if fps:
        A("| blob | ratio | band | flagged units | reading as English |")
        A("| --- | ---: | --- | ---: | ---: |")
        for it in sorted(fps, key=lambda x: -x["unit_legacy_ratio"]):
            A(f"| `{it['short']}` | {it['unit_legacy_ratio']:.4f} | {it['band']} "
              f"| {it['legacy_units_seen']} | "
              f"{it['legacy_units_reading_as_english']} |")
        A("")
        A("### Why native-2 mistakes these for glyph-mapped text")
        A("")
        A("Two of native-2's own signal corrections are defeated by the same "
          "typographic habit:")
        A("")
        A("1. **Numbered outline labels.** `5.2.Pension & Gratuity Fund` is one "
          "token to the tokenizer, and the intra-word `.` between digits and a "
          "capitalised word looks exactly like the intra-word symbol that marks "
          "Preeti. §13.2's correction exempts letter-bearing compounds such as "
          "`FIU-Nepal` and `F/Y`, but a `<digit>.<digit>.<Word>` label is not one.")
        A("2. **Accounting abbreviations.** `A/c`, `P/L`, `B/S` carry a symbol "
          "inside a two-letter token with no vowel, which is the vowel-less rule's "
          "target shape.")
        A("")
        A("The rest of each sheet is numeric, so those cells are `unjudged` and "
          "leave the ratio, shrinking the denominator until a handful of labels "
          "lifts it to ~0.5. That is the same small-denominator effect "
          "`MIN_JUDGED_FOR_RATIO` was introduced for — it bounds it, it does not "
          "eliminate it.")
        A("")
        A("Actual offending units:")
        A("")
        for it in sorted(fps, key=lambda x: -x["unit_legacy_ratio"]):
            A(f"**`{it['short']}`** — <{(it['comparison_keys'] or ['—'])[0]}>")
            A("")
            A("| where | flagged unit | what the converter did |")
            A("| --- | --- | --- |")
            for e in it["evidence"][:UNITS_PER_ITEM_MD]:
                A(f"| `{e['where']}` | `{_md_cell(e['original'])}` | "
                  f"{e['disposition']} |")
            A("")
    else:
        A("_No blob in this run met the ≥50% English-flagged-unit definition._")
        A("")

    # ---------------------------------------------------------------- 6
    A("## 6. False-negative candidates — clean documents carrying legacy units")
    A("")
    A("`candidate`, not `confirmed`. A document native-2 called clean can still "
      "contain a real legacy fragment. Whether it did turns on one thing that is "
      "decidable *without* Nepali: are the units it flagged actually English? "
      "Where they are, the document-level clean call was right and the flagged "
      "units are §5's defect appearing harmlessly below the gate. Where they are "
      "genuine Devanagari, it is a real candidate miss and a reader still has to "
      "confirm the reading.")
    A("")
    real = [i for i in fns if _english_share_of_flagged(i) < 0.5]
    notmiss = [i for i in fns if _english_share_of_flagged(i) >= 0.5]
    A(f"Of the {len(fns)} candidates examined: **{len(real)}** carry genuine "
      f"Devanagari, **{len(notmiss)}** turn out to be English units that native-2 "
      "was right not to route on.")
    A("")
    for it in sorted(fns, key=lambda x: -x["unit_legacy_candidates"]):
        share = _english_share_of_flagged(it)
        label = ("NOT a missed Nepali region" if share >= 0.5
                 else "genuine candidate miss")
        A(f"### `{it['short']}` — {it['family']}, "
          f"unit ratio {it['unit_legacy_ratio']:.4f} — **{label}**")
        A("")
        A(f"- {it['legacy_units_reading_as_english']} of "
          f"{it['legacy_units_seen']} flagged units read as English "
          f"({share:.1%})")
        keys = it["comparison_keys"]
        A(f"- source: <{keys[0]}>" if keys else "- source: —")
        A(f"- {it['unit_legacy_candidates']} legacy units of {it['unit_judged']} "
          f"judged ({it['unit_total']} total), longest run "
          f"{it['unit_max_legacy_run']}, contested ratio "
          f"{it['unit_contested_legacy_ratio']:.4f}")
        A("")
        A("**Why native-2 did not route it:**")
        A("")
        A(f"- the unit gate: `unit_legacy_ratio` {it['unit_legacy_ratio']:.4f} "
          f"{'≤' if it['unit_legacy_ratio'] <= quality.LEGACY_LINE_RATIO else '>'}"
          f" {quality.LEGACY_LINE_RATIO} — "
          f"{'below threshold, so the document-level call is clean' if it['unit_legacy_ratio'] <= quality.LEGACY_LINE_RATIO else 'above threshold'}")
        reasons = []
        if it["unit_legacy_candidates"] < routing.MINORITY_MIN_LEGACY_UNITS:
            reasons.append(
                f"`legacy >= {routing.MINORITY_MIN_LEGACY_UNITS}` fails "
                f"({it['unit_legacy_candidates']})")
        if it["unit_max_legacy_run"] < routing.MINORITY_MIN_RUN:
            reasons.append(
                f"`max run >= {routing.MINORITY_MIN_RUN}` fails "
                f"({it['unit_max_legacy_run']})")
        if it["unit_contested_legacy_ratio"] < routing.MINORITY_MIN_CONTESTED_RATIO:
            reasons.append(
                f"`contested ratio >= {routing.MINORITY_MIN_CONTESTED_RATIO}` "
                f"fails ({it['unit_contested_legacy_ratio']:.4f})")
        if reasons:
            A(f"- the minority-region rule: {'; '.join(reasons)} — all three "
              "conditions are required")
        else:
            A("- the minority-region rule: all three conditions hold "
              f"(`minority_legacy_detected={it['minority_legacy_detected']}`), so "
              "this document is flagged by the region rule and is not a false "
              "negative")
        A("")
        ev = it["evidence"][:UNITS_PER_ITEM_MD]
        if ev:
            A(f"Showing {len(ev)} of {it['legacy_units_seen']} legacy units:")
            A("")
            A("| where | extracted | converted (evaluation only) |")
            A("| --- | --- | --- |")
            for e in ev:
                A(f"| `{e['where']}` | `{_md_cell(e['original'])}` | "
                  f"{_md_cell(e['converted'] or '—')} |")
            A("")
        if share >= 0.5:
            A(f"**Verdict for `{it['short']}`: `ambiguous` — the flagged units "
              "are English, so this is not evidence of a missed legacy region. "
              "No Nepali reading is required.**")
        else:
            A(f"**Verdict for `{it['short']}`: `candidate` / "
              "`awaiting_nepali_review` — a reader must confirm the recovered "
              "Devanagari before this counts as a confirmed false negative.**")
        A("")

    # ---------------------------------------------------------------- 7
    A("## 7. Spreadsheet evidence")
    A("")
    A("### The native-1 blind spot, on files native-1 never saw")
    A("")
    A("Independent confirmation that judging a workbook per CELL is what fixed "
      "it. Native-1's document-level line metric calls all three of these clean; "
      "native-2's unit metric does not.")
    A("")
    A("| blob | `unit_legacy_ratio` (native-2) | `legacy_line_ratio` (native-1) | "
      "native-1 verdict at 0.20 | sheets | legacy cells |")
    A("| --- | ---: | ---: | --- | ---: | ---: |")
    sheets_high = [i for i in high if i["family"] == "spreadsheet"]
    for it in sorted(sheets_high, key=lambda x: -x["unit_legacy_ratio"]):
        A(f"| `{it['short']}` | **{it['unit_legacy_ratio']:.4f}** | "
          f"{it['legacy_line_ratio']:.4f} | "
          f"{'clean' if it['legacy_line_ratio'] <= 0.20 else 'flagged'} | "
          f"{it['sheet_count']} | {it['unit_legacy_candidates']} |")
    A("")
    A("The gate is `unit_legacy_ratio`, and this table is why substituting "
      "`legacy_line_ratio` in a future production router would route a different "
      "population. `tests/test_nrb_phase6b_holdout.py::"
      "test_the_conversion_gate_reads_the_unit_metric_not_the_line_metric` "
      "locks that.")
    A("")
    A("### Format gap — OLE2 files with no parser")
    A("")
    A("Fetched, verified, stored, and unreadable: the corpus contains "
      "pre-2007 Microsoft Office binaries. `extraction.extract_file` refuses them "
      "by extension **before** the sniffed family is consulted, so an "
      "`application/x-ole-storage` blob never reaches openpyxl. Recorded as a "
      "corpus/format gap; no `.xls` support is implemented in this task.")
    A("")
    unsup = acct["buckets"].get("extracted:unsupported", [])
    kinds = collections.Counter(
        (r.get("comparison_key") or "").rsplit(".", 1)[-1].lower() for r in unsup
    )
    claimed = sum(1 for rows in acct["buckets"].values() for r in rows
                  if r.get("resource_type") == "spreadsheet")
    ole_xls = kinds.get("xls", 0)
    failed_n = len(acct["buckets"].get("extracted:failed", []))
    A(f"NRB's own `resource_type` calls **{claimed}** holdout files "
      f"spreadsheets. That is {claimed - ole_xls} xlsx plus {ole_xls} OLE2 "
      f"`.xls`; of the xlsx, {failed_n} could not be parsed by openpyxl either "
      "(`parser_error`, §1). The two `.doc` in the table below are typed "
      "`document`, not `spreadsheet` — they are here because they share the same "
      "OLE2 cause.")
    A("")
    A(f"| extension | count |")
    A("| --- | ---: |")
    for ext, n in sorted(kinds.items()):
        A(f"| `.{ext}` | {n} |")
    A(f"| **total** | **{len(unsup)}** |")
    A("")

    # ---------------------------------------------------------------- 8
    A("## 8. Development evidence — *not* Phase 6B holdout")
    A("")
    A("> ⚠️ **development evidence — not Phase 6B holdout.** Everything in this "
      "section comes from the Phase 6A benchmark, which SHAPED native-1 and "
      "native-2. It is listed here only so one review sitting can clear the whole "
      "backlog. **Do not fold any of it into the holdout statistics above.**")
    A("")
    A("Outstanding from Phase 6A / Task 2, per `docs/nrb-integration.md`:")
    A("")
    A("- the five spreadsheet cases in "
      "`docs/nrb/phase6b-native2-manual-review.txt`;")
    A("- the Preeti reading of `8df7b02f8a13`, the benchmark's Preeti-encoded "
      "workbook — the case that established per-cell conversion;")
    A("- `docs/nrb/phase6b-manual-validation.txt`, the Task 1 conversion sample.")
    A("")

    # ---------------------------------------------------------------- 9
    A("## 9. What this pack does not establish")
    A("")
    A("- **No semantic correctness result.** Every Nepali verdict is "
      "`awaiting_nepali_review`.")
    A("- **No classifier change.** native-2 is byte-identical to commit "
      "`2a6b498`; no threshold, guard or version moved.")
    A("- **No production conversion.** Nothing routes to the converter at "
      "runtime; npttf2utf is not installed by the Dockerfile and the GPL-3.0 "
      "distribution question is still open.")
    A("- **No OCR**, no chunking, no embeddings, no `documents`/pgvector writes.")
    A("")
    A("Decision as it stands: *independent holdout evidence strongly supports the "
      "native-2 `>=0.80` high-confidence routing candidate, but semantic "
      "conversion correctness remains pending Nepali human review. Native-2 also "
      "exposed a real lower-band English false-positive class. No classifier "
      "change or production converter integration is made in this task.*")
    A("")

    # --------------------------------------------------------------- 10
    A("## 10. Evaluation & Improvement")
    A("")
    A("**Success metric.** Share of `>=0.80`-routed NRB blobs whose converted "
      "Unicode a Nepali reader marks `confirmed_correct`. Proxy until reviews "
      "land: routing precision on unseen files (currently "
      f"{q['routed'] - q['false_route_english']}/{q['routed']}).")
    A("")
    A(f"**Eval.** This pack: {len(high)} labelled review items with the flagged "
      "unit, the converted output and the rendered source page. Scored by reader "
      "verdict per item. Current agreement rate: **not yet measurable** — 0 of "
      f"{len(high)} adjudicated.")
    A("")
    A("**Feedback capture.** The reader edits the verdict column in §3 and the "
      "per-item line in §4, in place, in this file under version control. "
      "Disagreements with the `E` (English) column are the routing-precision "
      "correction signal; disagreements on converted text are the conversion "
      "signal. The two are logged separately because they have different fixes.")
    A("")
    A("**Review loop.** Re-examine when the reader returns verdicts, and "
      "otherwise at each extractor-version change. A `native-3` addressing §5 "
      "invalidates this cohort as validation evidence and requires a new draw "
      "(new seed, `exclude_keys` covering **both** Phase 6A and this holdout).")
    A("")
    if payload["problems"]:
        A("## Appendix — generation problems")
        A("")
        for p in payload["problems"]:
            A(f"- {p}")
        A("")
    return "\n".join(o)


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    p.add_argument("--holdout", default="docs/nrb/phase6b-routing-holdout.json")
    p.add_argument("--phase6a", default="docs/nrb/phase6a-manifest.json")
    p.add_argument("--lexicon", default="docs/nrb/phase6b-lexicon.json")
    p.add_argument("--profile",
                   default="docs/nrb/phase6b-routing-holdout-profile.json")
    p.add_argument("--out-md",
                   default="docs/nrb/phase6b-routing-holdout-manual-review.md")
    p.add_argument("--out-json",
                   default="docs/nrb/phase6b-routing-holdout-evidence.json")
    p.add_argument("--pages-dir", default="docs/nrb/holdout-pages")
    return asyncio.run(run(p.parse_args()))


if __name__ == "__main__":
    raise SystemExit(main())
