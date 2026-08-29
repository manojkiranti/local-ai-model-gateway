"""Local tool: edit_excel — apply ONE edit to an uploaded .xlsx, as a NEW file.

Owner-scoped by file_id (see files/source.py). The uploaded workbook is never
modified: the edit is applied in memory and saved as a new `generated_files`
entry, so a wrong edit costs nothing and the original stays intact.

**Loads `data_only=False`.** `readers.py` deliberately loads `data_only=True`
so it can never evaluate a formula, but on this WRITE path that setting
destroys them: openpyxl 3.1.5 replaces every formula with its last cached value
on save, so a workbook a human saved from Excel comes back with the right
numbers and no live model behind them — correct-looking and silently wrong.
`tests/test_edit_excel_tool.py::test_a_formula_elsewhere_in_the_workbook_survives_the_edit`
is the regression guard.

ONE operation per call, deliberately. A single ordered list of edits reads as
convenient and is a correctness trap: deleting a row shifts every A1 reference
after it, and the model cannot reliably predict the shift. Each operation is
batched internally instead, so one call still does one whole job.
"""

from __future__ import annotations

import asyncio
from io import BytesIO
from pathlib import Path
from typing import Any

from ...files import editor
from ...files.store import XLSX_MEDIA_TYPE, file_store, resolve_file
from .base import LocalToolSpec

# Bound on how big a workbook we will hold in memory. openpyxl's editable mode
# is not read-only/streaming, so the whole book is resident. Checked with a
# cheap read-only pass BEFORE the real load — the `images.MAX_IMAGE_PIXELS`
# rule: refuse from the declared dimensions, don't find out by exhausting RAM.
MAX_EDIT_CELLS = 1_000_000

# Which keys each operation takes. Keeps the arg plumbing declarative and stops
# an unrelated key being forwarded into the editor as a surprise kwarg.
_OP_PARAMS: dict[str, tuple[str, ...]] = {
    "set_cells": ("cells",),
    "append_rows": ("rows",),
    "add_column": ("header", "values"),
    "delete_rows": ("rows",),
    "delete_columns": ("columns",),
}


def _count_cells(path: Path) -> int:
    """Cell count across every sheet, from a cheap read-only open."""
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True)
    try:
        return sum((ws.max_row or 0) * (ws.max_column or 0) for ws in wb.worksheets)
    finally:
        wb.close()


def _apply(path: Path, operation: str, sheet: str | None, params: dict[str, Any]) -> tuple[bytes, editor.EditResult]:
    """Load, edit and re-serialize. Sync (openpyxl) — run in a thread."""
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=False)  # see the module docstring
    try:
        result = editor.apply(wb, operation=operation, sheet=sheet, **params)
        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue(), result
    finally:
        wb.close()


async def _edit_excel(args: dict[str, Any]) -> str:
    file_id = args.get("file_id")
    if not isinstance(file_id, str) or not file_id.strip():
        return "ERROR: 'file_id' is required (the id of an uploaded spreadsheet)."
    record = await resolve_file(file_id.strip())
    if record is None:
        return "ERROR: no such file (unknown id, or you don't own it)."

    operation = args.get("operation")
    if operation not in _OP_PARAMS:
        return (
            f"ERROR: unknown operation {operation!r}. Valid operations are: "
            f"{', '.join(sorted(_OP_PARAMS))}."
        )

    path = Path(record.path)
    suffix = path.suffix.lower()
    if suffix != ".xlsx":
        # Name the extension the caller actually has. Reporting every non-.xlsx
        # file as a .csv sent the model to read_excel, which then fails on a PDF
        # and loops.
        if suffix == ".csv":
            fix = (
                "Read it with read_excel and write a new file with create_excel."
            )
        else:
            fix = (
                "Only spreadsheets can be edited — read this one with "
                "read_document (.pdf/.docx/.txt/.md/.json) or read_image."
            )
        return (
            f"ERROR: edit_excel works on .xlsx workbooks only, and "
            f"'{record.filename}' is a {suffix or 'file with no extension'}. {fix}"
        )

    try:
        cells = await asyncio.to_thread(_count_cells, path)
    except Exception as exc:  # noqa: BLE001 - a corrupt/unreadable book, not a crash
        return f"ERROR: could not read the spreadsheet ({exc})."
    if cells > MAX_EDIT_CELLS:
        return (
            f"ERROR: this workbook is too large to edit ({cells:,} cells; the limit "
            f"is {MAX_EDIT_CELLS:,}). Narrow it down and use create_excel instead."
        )

    sheet = args.get("sheet")
    sheet = str(sheet) if sheet is not None else None
    # Every operand key is passed through even when absent, so the editor's own
    # validation produces the message. Dropping absent keys instead turned a
    # missing operand into a raw TypeError naming an internal function.
    params = {k: args.get(k) for k in _OP_PARAMS[operation]}

    try:
        data, result = await asyncio.to_thread(_apply, path, operation, sheet, params)
    except editor.EditError as exc:
        return f"ERROR: {exc}."
    except Exception as exc:  # noqa: BLE001 - report back, don't raise into the loop
        return f"ERROR: could not edit the spreadsheet ({exc})."

    stem = Path(record.filename).stem or "book"
    filename = str(args.get("filename") or f"{stem} (edited).xlsx")
    if not filename.lower().endswith(".xlsx"):
        filename += ".xlsx"

    new = await file_store.save(data, filename=filename, media_type=XLSX_MEDIA_TYPE)
    return (
        f"Edited '{record.filename}' → new file '{new.filename}' ({new.size} bytes). "
        f"Sheet '{result.sheet_name}': {result.detail}. "
        f"The original file is unchanged. "
        f"Download the edited file at: GET /v1/files/{new.id} "
        f"(to make further edits, call edit_excel again with file_id={new.id})"
    )


SPEC = LocalToolSpec(
    name="edit_excel",
    description=(
        "Modify an uploaded .xlsx spreadsheet and return a NEW file — the "
        "original is never changed. Give the 'file_id', one 'operation', and "
        "that operation's arguments: 'set_cells' (cells=[{cell:'B2',value:…}]), "
        "'append_rows' (rows=[[…]]), 'add_column' (header + one value per data "
        "row), 'delete_rows' (rows=[2,5] — 1-based DATA row numbers as read_excel "
        "shows them, header excluded), 'delete_columns' (columns=['qty']). ONE "
        "operation per call; call again with the returned file_id to make another "
        "change. Use inspect_excel/read_excel first to see the sheets, headers and "
        "row numbers. Values are literal — this tool does not write Excel formulas."
    ),
    parameters={
        "type": "object",
        "properties": {
            "file_id": {"type": "string", "description": "Id of an uploaded .xlsx workbook."},
            "operation": {
                "type": "string",
                "enum": ["set_cells", "append_rows", "add_column", "delete_rows", "delete_columns"],
                "description": "The single edit to apply.",
            },
            "sheet": {"type": "string", "description": "Sheet name or 1-based index (default: first sheet)."},
            "cells": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "cell": {"type": "string", "description": "Single A1 reference, e.g. 'B2'."},
                        "value": {"description": "Literal value to write."},
                    },
                    "required": ["cell", "value"],
                },
                "description": "set_cells: the cells to write.",
            },
            "rows": {
                "type": "array",
                "description": (
                    "append_rows: rows of cell values to add. "
                    "delete_rows: 1-based DATA row numbers to remove."
                ),
            },
            "header": {"type": "string", "description": "add_column: the new column's header."},
            "values": {
                "type": "array",
                "description": "add_column: one value per existing data row.",
            },
            "columns": {
                "type": "array",
                "items": {"type": "string"},
                "description": "delete_columns: header names to remove.",
            },
            "filename": {"type": "string", "description": "Optional name for the new file."},
        },
        "required": ["file_id", "operation"],
    },
    func=_edit_excel,
)
