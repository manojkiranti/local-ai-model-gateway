"""Apply ONE structural edit to an openpyxl workbook.

Pure module — no DB, no HTTP, no file store — for the reason `aggregate.py` and
`readers.py` are pure: the rules that decide what happens to a user's data
should be provable without standing anything up. The tool layer
(`app/tools/local/edit_excel.py`) owns resolving a file_id, the cell-count cap
and saving the result as a NEW file; this module owns only the edit itself.

**The workbook MUST be loaded `data_only=False`.** Measured on openpyxl 3.1.5:
loading `data_only=True` and saving replaces every formula with its cached
value, so `=B2*C2` becomes the number it last evaluated to — or nothing at all
if the file was never opened in Excel. `readers.py` uses `data_only=True`
deliberately (it must never evaluate a formula), and reusing that setting on
this WRITE path silently converts a live spreadsheet into static numbers while
every signal still reports success. Never "unify" the two.

Writing a formula is likewise refused rather than escaped: a value beginning
with `=` becomes a live formula in Excel, and our own read path
(`data_only=True`) would then show it as blank.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from typing import Any

from openpyxl.utils import column_index_from_string, get_column_letter

# A single-cell A1 reference: up to 3 column letters, then a 1-based row.
# Ranges ("B2:C3"), absolute refs ("$B$2") and sheet-qualified refs are refused
# on purpose — every operation here names cells one at a time.
_A1 = re.compile(r"^([A-Za-z]{1,3})([1-9][0-9]*)$")

# Excel's real grid. The regex alone allows up to 'ZZZ' (18278) and any row, and
# openpyxl writes those without complaint — producing a file outside the format's
# limits whose inflated dimension then trips this tool's own cell cap on every
# later edit. Bound them here.
MAX_COLUMN = 16384      # XFD
MAX_ROW = 1_048_576


class EditError(Exception):
    """The requested edit is not something we will do (bad ref, formula, …)."""


@dataclass
class EditResult:
    """What an applied operation actually did, for the model to read back."""

    operation: str
    sheet_name: str
    detail: str


def _check_value(value: Any) -> Any:
    if isinstance(value, str) and value.startswith("="):
        raise EditError(
            f"refusing to write {value!r}: a leading '=' would create a live Excel "
            "formula, and this tool writes literal values only"
        )
    return value


def _check_ref(ref: Any) -> str:
    match = _A1.match(ref.strip()) if isinstance(ref, str) else None
    if match is None:
        raise EditError(
            f"invalid cell reference {ref!r}: give a single cell like 'B2' "
            "(not a range, not an absolute or sheet-qualified reference)"
        )
    letters, digits = match.groups()
    if column_index_from_string(letters.upper()) > MAX_COLUMN or int(digits) > MAX_ROW:
        raise EditError(
            f"invalid cell reference {ref!r}: outside Excel's grid "
            f"(last cell is XFD{MAX_ROW})"
        )
    return ref.strip().upper()


def _listing(items: list[str], limit: int = 10) -> str:
    """Name at most `limit` items, then say how many more there are.

    The result is read back into the model's context, where `agent/loop.py`
    truncates from the END at MAX_TOOL_RESULT_CHARS — a 400-entry list would
    spend that budget on something nobody reads.
    """
    if len(items) <= limit:
        return ", ".join(items)
    return f"{', '.join(items[:limit])} … and {len(items) - limit} more"


def _set_cells(ws, cells: Any) -> EditResult:
    if not isinstance(cells, list) or not cells:
        raise EditError("'cells' must be a non-empty array of {cell, value} objects")

    # Validate the WHOLE batch before writing anything: a bad entry late in the
    # list must not leave the earlier ones applied, or the caller cannot tell
    # what state the sheet is in.
    checked: list[tuple[str, Any]] = []
    for entry in cells:
        if not isinstance(entry, dict):
            raise EditError("each entry of 'cells' must be an object {cell, value}")
        ref = _check_ref(entry.get("cell"))
        # An ABSENT 'value' used to arrive as None and silently blank the cell
        # while reporting success. An EXPLICIT null still clears it — that is a
        # deliberate edit, and refusing it would remove the only way to empty a
        # cell.
        if "value" not in entry:
            raise EditError(
                f"cell {ref} has no 'value': give {{\"cell\": \"{ref}\", \"value\": …}}, "
                "or an explicit null to clear it"
            )
        checked.append((ref, _check_value(entry["value"])))

    for ref, value in checked:
        ws[ref] = value
    return EditResult(
        operation="set_cells",
        sheet_name=ws.title,
        detail=f"set {len(checked)} cell(s): {_listing([ref for ref, _ in checked])}",
    )


def _append_rows(ws, rows: Any) -> EditResult:
    if not isinstance(rows, list) or not rows:
        raise EditError("'rows' must be a non-empty array of row arrays")

    checked: list[list[Any]] = []
    for row in rows:
        if not isinstance(row, (list, tuple)):
            raise EditError("each entry of 'rows' must be an array of cell values")
        checked.append([_check_value(v) for v in row])

    # Not ws.append(): it starts from ws.max_row, so a styled empty row far
    # below the data pushes the new record past the gap instead of onto the
    # first free line.
    next_row = _last_populated_row(ws) + 1
    for offset, row in enumerate(checked):
        for column, value in enumerate(row, start=1):
            ws.cell(row=next_row + offset, column=column, value=value)
    return EditResult(
        operation="append_rows",
        sheet_name=ws.title,
        detail=f"appended {len(checked)} row(s); the sheet now has "
        f"{_data_row_count(ws)} data row(s)",
    )


def _add_column(ws, header: Any = None, values: Any = None) -> EditResult:
    if not isinstance(header, str) or not header.strip():
        raise EditError("'header' must be a non-empty column-name string")
    if not isinstance(values, list):
        raise EditError("'values' must be an array of cell values, one per data row")

    expected = _data_row_count(ws)
    if len(values) != expected:
        raise EditError(
            f"'values' has {len(values)} entr(ies) but the sheet has {expected} "
            "data row(s) — supply exactly one value per data row"
        )

    checked = [_check_value(v) for v in values]
    _check_value(header)
    # openpyxl reports max_column == 1 for a BLANK sheet, so "max_column + 1"
    # would skip column A entirely.
    col = 1 if _is_blank(ws) else ws.max_column + 1
    ws.cell(row=1, column=col, value=header.strip())
    for offset, value in enumerate(checked, start=2):
        ws.cell(row=offset, column=col, value=value)
    return EditResult(
        operation="add_column",
        sheet_name=ws.title,
        detail=f"added column '{header.strip()}' as column "
        f"{get_column_letter(col)} with {len(checked)} value(s)",
    )


def _is_blank(ws) -> bool:
    """True for a sheet with nothing in it (openpyxl still reports 1x1)."""
    return ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None


def _formula_text(cell) -> str | None:
    """The formula in a cell, however openpyxl chose to represent it.

    A plain formula comes back as a `str`, but an ARRAY/CSE formula (and a data
    table) comes back as an `ArrayFormula`/`DataTableFormula` OBJECT. An
    isinstance(str) test walks straight past those, which let the guard below
    approve exactly the corruption it exists to prevent. `data_type == "f"` is
    the one check that covers every representation.
    """
    if cell.data_type != "f":
        return None
    value = cell.value
    if isinstance(value, str):
        return value
    # ArrayFormula/DataTableFormula carry the source in `.text`.
    return getattr(value, "text", None) or ""


def _first_formula(ws) -> str | None:
    for row in ws.iter_rows():
        for cell in row:
            if _formula_text(cell) is not None:
                return cell.coordinate
    return None


def _guard_shifting_formulas(wb, ws) -> None:
    """Refuse a row/column deletion on a workbook that contains formulas.

    openpyxl's `delete_rows`/`delete_cols` are NOT formula-aware: Excel rewrites
    every reference when a row is removed, openpyxl leaves the text exactly as
    it was. Measured — deleting a row above `=SUM(B2:B4)` leaves it spelled
    `=SUM(B2:B4)` while the rows it names have moved, so it silently sums the
    wrong records. Nothing errors and the file opens fine, which is the §18
    failure class again. `data_only=False` protects formulas from being
    OVERWRITTEN; it cannot make a shift correct, so these two operations refuse
    instead of returning a workbook whose formulas quietly mean something else.
    """
    where = _first_formula(ws)
    if where is not None:
        raise EditError(
            f"refusing to delete from sheet '{ws.title}': it contains a formula "
            f"({where}) and deleting shifts the rows/columns its references point "
            "at without rewriting them, so the formula would silently compute the "
            "wrong answer. Overwrite the formula cells with set_cells first, or "
            "download the file and delete the rows in Excel, which rewrites the "
            "references properly. Do NOT rebuild the file with create_excel from "
            "a read_excel dump — that output is capped and would drop rows"
        )
    for other in wb.worksheets:
        if other is ws:
            continue
        for row in other.iter_rows():
            for cell in row:
                text = _formula_text(cell)
                if text is not None and ws.title in text:
                    raise EditError(
                        f"refusing to delete from sheet '{ws.title}': sheet "
                        f"'{other.title}' has a formula ({cell.coordinate}) that "
                        "references it, and deleting would shift what that formula "
                        "points at without rewriting it"
                    )


def _last_populated_row(ws) -> int:
    """The last row holding a VALUE, ignoring rows that carry only formatting.

    `ws.max_row` counts a row styled but never filled — ordinary in real
    uploads — while `readers._xlsx_sheet_grid` trims trailing empty rows. The
    two disagreeing meant `read_excel` showed 3 data rows while the editor
    believed 9: `add_column` then demanded nine values for a three-row sheet,
    `delete_rows` accepted numbers the read tools never displayed, and
    `append_rows` landed past the gap. The read tools are what the model saw,
    so this side is the one that has to move.
    """
    for row in range(ws.max_row, 0, -1):
        if any(cell.value is not None for cell in ws[row]):
            return row
    return 0


def _data_row_count(ws) -> int:
    """Rows below the header row. An empty sheet has none."""
    return max(0, _last_populated_row(ws) - 1)


def _delete_rows(ws, rows: Any) -> EditResult:
    """Delete DATA rows by their 1-based number as `read_excel` displays them.

    Every number is resolved against the numbering the caller saw, then applied
    in DESCENDING sheet order — deleting top-down would shift each later target
    up by one and silently remove the wrong record.
    """
    targets = _int_list(rows, field="rows")
    available = _data_row_count(ws)
    bad = [n for n in targets if n < 1 or n > available]
    if bad:
        raise EditError(
            f"row number(s) {', '.join(str(n) for n in bad)} are out of range: "
            f"the sheet has {available} data row(s), numbered 1..{available}"
        )

    # +1 because data row N is sheet row N+1 (row 1 is the header).
    for n in sorted(set(targets), reverse=True):
        ws.delete_rows(n + 1)
    return EditResult(
        operation="delete_rows",
        sheet_name=ws.title,
        detail=f"deleted {len(set(targets))} data row(s) "
        f"({_listing([str(n) for n in sorted(set(targets))])}); "
        f"{_data_row_count(ws)} row(s) remain",
    )


def _delete_columns(ws, columns: Any) -> EditResult:
    if not isinstance(columns, list) or not columns:
        raise EditError("'columns' must be a non-empty array of header names")

    headers = [c.value for c in ws[1]] if ws.max_row else []
    # Resolve first, then dedupe on the INDEX: 'qty' and 'B' can name one
    # column, and deleting it twice destroys the column that shifted into its
    # place. Keyed by the caller's string, that read as two successful deletes.
    resolved: dict[int, str] = {}
    for name in columns:
        resolved.setdefault(_column_index(headers, name), str(name))

    for idx in sorted(resolved, reverse=True):
        ws.delete_cols(idx)
    return EditResult(
        operation="delete_columns",
        sheet_name=ws.title,
        detail=f"deleted {len(resolved)} column(s): "
        f"{_listing([resolved[i] for i in sorted(resolved)])}",
    )


def _column_index(headers: list[Any], name: Any) -> int:
    """1-based column index for a header NAME, falling back to a column letter.

    Header name wins when both could match, because that is what the model read
    out of `inspect_excel`; a bare letter is the fallback for an unheaded column.
    """
    target = str(name).strip()
    for i, header in enumerate(headers, start=1):
        if header is not None and str(header).strip().casefold() == target.casefold():
            return i
    if target.isalpha() and len(target) <= 3:
        try:
            index = column_index_from_string(target.upper())
        except ValueError:
            index = None
        # Bound it to the sheet: delete_cols(26) on a 3-column sheet is a
        # silent no-op that used to be reported as a successful delete.
        if index is not None and 1 <= index <= len(headers):
            return index
    known = ", ".join(str(h) for h in headers if h is not None) or "(none)"
    raise EditError(f"no column named '{name}' (headers are: {known})")


def _int_list(value: Any, *, field: str) -> list[int]:
    if not isinstance(value, list) or not value:
        raise EditError(f"'{field}' must be a non-empty array of integers")
    out: list[int] = []
    for item in value:
        if isinstance(item, bool) or not isinstance(item, int):
            raise EditError(f"'{field}' must contain integers only (got {item!r})")
        out.append(item)
    return out


_OPERATIONS = {
    "set_cells": _set_cells,
    "append_rows": _append_rows,
    "add_column": _add_column,
    "delete_rows": _delete_rows,
    "delete_columns": _delete_columns,
}

# Operations that MOVE existing cells, and so can invalidate a formula's
# references without openpyxl rewriting them. See `_guard_shifting_formulas`.
_SHIFTING_OPERATIONS = frozenset({"delete_rows", "delete_columns"})


def apply(wb, *, operation: str, sheet: str | None = None, **params: Any) -> EditResult:
    """Apply one named operation to one sheet of `wb`, in place."""
    if operation not in _OPERATIONS:
        raise EditError(
            f"unknown operation {operation!r} (expected one of: "
            f"{', '.join(sorted(_OPERATIONS))})"
        )
    ws = _resolve_sheet(wb, sheet)
    if operation in _SHIFTING_OPERATIONS:
        _guard_shifting_formulas(wb, ws)
    return _OPERATIONS[operation](ws, **params)


def _resolve_sheet(wb, sheet: str | None):
    """Same addressing rule as `readers._resolve_sheet`: name, or 1-based index."""
    names = list(wb.sheetnames)
    if not sheet:
        return wb[names[0]]
    target = str(sheet).strip()
    if target.isdigit():
        idx = int(target)
        if 1 <= idx <= len(names):
            return wb[names[idx - 1]]
        raise EditError(f"sheet index {idx} out of range (1..{len(names)})")
    for name in names:
        if name.strip().casefold() == target.casefold():
            return wb[name]
    raise EditError(f"no sheet named '{sheet}' (have: {', '.join(names)})")
