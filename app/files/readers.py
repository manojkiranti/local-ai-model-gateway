"""Normalize an uploaded spreadsheet into a plain rows/headers `Table`.

Pure module — no DB, no HTTP. Both the upload route (for its parse check +
summary) and the read tools (`inspect_excel`/`read_excel`) go through here, so
`.xlsx` and `.csv` behave identically downstream.

Safety choices:
  * xlsx is opened `read_only=True, data_only=True` — we read the LAST CACHED
    value of any formula cell and NEVER evaluate a formula (a workbook never
    opened in Excel simply has empty formula cells; that's fine).
  * every cell is coerced to a clipped string, and both a row cap and a total
    character cap bound how much can ever reach the model's context.
  * a `.csv` is presented as ONE pseudo-sheet named after the file stem, so the
    tools don't need format-specific branches.
"""

from __future__ import annotations

import csv as _csv
from contextlib import contextmanager
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable, Iterator, Optional

# Caps (shared with the read tool). Deliberately small: the agent context is
# ~16k tokens, so dumping thousands of rows would blow it — better to truncate
# loudly and let the model page or filter.
READ_MAX_ROWS = 200
READ_MAX_CHARS = 40_000
CELL_MAX_CHARS = 500
SAMPLE_ROWS = 10


class ReadError(Exception):
    """The file could not be parsed as a spreadsheet (corrupt/unsupported)."""


class SheetNotFound(Exception):
    """A named sheet does not exist in the workbook."""


@dataclass
class SheetInfo:
    sheet_name: str
    total_rows: int          # data rows (excludes the header row)
    total_cols: int
    headers: list[str]
    sample_rows: list[list[str]] = field(default_factory=list)
    hidden: bool = False


@dataclass
class Table:
    sheet_name: str
    headers: list[str]
    rows: list[list[str]]
    total_rows: int          # total data rows available in the sheet
    total_cols: int
    truncated: bool          # True if fewer rows/chars returned than exist
    all_sheets: list[str] = field(default_factory=list)  # names of every sheet


# --------------------------------------------------------------------------- #
# Cell coercion
# --------------------------------------------------------------------------- #
def _cell(value: object) -> str:
    """Coerce any cell to a clipped display string. None -> ''."""
    if value is None:
        return ""
    text = str(value)
    if len(text) > CELL_MAX_CHARS:
        text = text[:CELL_MAX_CHARS] + "…"
    return text


def _is_xlsx(path: Path) -> bool:
    return path.suffix.lower() in (".xlsx",)


# --------------------------------------------------------------------------- #
# Raw grid loaders (return list-of-rows, header included as row 0)
# --------------------------------------------------------------------------- #
def _xlsx_sheet_grid(ws, *, limit: Optional[int] = None) -> list[list[str]]:
    grid: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        grid.append([_cell(v) for v in row])
        if limit is not None and len(grid) >= limit:
            break
    # openpyxl read-only can yield trailing all-empty rows; trim them.
    while grid and all(c == "" for c in grid[-1]):
        grid.pop()
    return grid


def _csv_grid(path: Path, *, limit: Optional[int] = None) -> list[list[str]]:
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            sample = fh.read(8192)
            fh.seek(0)
            try:
                dialect = _csv.Sniffer().sniff(sample, delimiters=",;\t|")
            except _csv.Error:
                dialect = _csv.excel  # default to comma
            reader = _csv.reader(fh, dialect)
            grid: list[list[str]] = []
            for row in reader:
                grid.append([_cell(v) for v in row])
                if limit is not None and len(grid) >= limit:
                    break
            return grid
    except (OSError, UnicodeError) as exc:
        raise ReadError(f"could not read CSV: {exc}") from exc


# --------------------------------------------------------------------------- #
# Workbook helpers
# --------------------------------------------------------------------------- #
def _open_xlsx(path: Path):
    try:
        from openpyxl import load_workbook

        return load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - openpyxl raises many types on bad input
        raise ReadError(f"could not read spreadsheet: {exc}") from exc


def _norm(name: str) -> str:
    return name.strip().lower()


# --------------------------------------------------------------------------- #
# Public: inspect every sheet
# --------------------------------------------------------------------------- #
def inspect_workbook(path: Path) -> list[SheetInfo]:
    """Structure of every sheet: headers, counts, a small sample. CSV -> one."""
    path = Path(path)
    if not _is_xlsx(path):
        grid = _csv_grid(path)
        return [_sheet_info_from_grid(path.stem, grid, hidden=False)]

    wb = _open_xlsx(path)
    try:
        infos: list[SheetInfo] = []
        for name in wb.sheetnames:
            ws = wb[name]
            # Chart-only / empty sheets have no usable cells — represent as empty.
            grid = _xlsx_sheet_grid(ws, limit=SAMPLE_ROWS + 1)
            total_rows = max((ws.max_row or 0) - 1, 0)  # minus header
            total_cols = ws.max_column or (len(grid[0]) if grid else 0)
            hidden = getattr(ws, "sheet_state", "visible") != "visible"
            info = _sheet_info_from_grid(
                name, grid, hidden=hidden,
                total_rows_override=total_rows, total_cols_override=total_cols,
            )
            infos.append(info)
        return infos
    finally:
        wb.close()


def _sheet_info_from_grid(
    name: str,
    grid: list[list[str]],
    *,
    hidden: bool,
    total_rows_override: Optional[int] = None,
    total_cols_override: Optional[int] = None,
) -> SheetInfo:
    headers = grid[0] if grid else []
    data = grid[1:]
    total_rows = total_rows_override if total_rows_override is not None else len(data)
    total_cols = total_cols_override if total_cols_override is not None else len(headers)
    return SheetInfo(
        sheet_name=name,
        total_rows=total_rows,
        total_cols=total_cols,
        headers=headers,
        sample_rows=data[:SAMPLE_ROWS],
        hidden=hidden,
    )


# --------------------------------------------------------------------------- #
# Public: load one sheet's rows (capped, pageable, projectable)
# --------------------------------------------------------------------------- #
def load_table(
    path: Path,
    *,
    sheet: Optional[str] = None,
    columns: Optional[list[str]] = None,
    start_row: int = 1,
    max_rows: Optional[int] = None,
) -> Table:
    """One sheet as a capped `Table`. `start_row` is 1-based into DATA rows
    (after the header). `columns` projects/reorders by header name."""
    path = Path(path)
    start_row = max(1, int(start_row or 1))
    cap = READ_MAX_ROWS if max_rows is None else min(int(max_rows), READ_MAX_ROWS)

    if not _is_xlsx(path):
        grid = _csv_grid(path)
        all_sheets = [path.stem]
        sheet_name = path.stem
    else:
        wb = _open_xlsx(path)
        try:
            all_sheets = list(wb.sheetnames)
            sheet_name = _resolve_sheet(all_sheets, sheet)
            grid = _xlsx_sheet_grid(wb[sheet_name])
        finally:
            wb.close()

    headers = grid[0] if grid else []
    data = grid[1:]
    total_rows = len(data)

    # Page window into the data rows.
    window = data[start_row - 1 : start_row - 1 + cap]
    truncated = (start_row - 1 + len(window)) < total_rows

    headers, window = _project(headers, window, columns)

    # Character cap: trim whole rows off the end until under the budget.
    truncated = _apply_char_cap(window, truncated)

    return Table(
        sheet_name=sheet_name,
        headers=headers,
        rows=window,
        total_rows=total_rows,
        total_cols=len(headers),
        truncated=truncated,
        all_sheets=all_sheets,
    )


def _resolve_sheet(sheetnames: list[str], sheet: Optional[str]) -> str:
    if not sheet:
        return sheetnames[0]
    target = str(sheet).strip()
    # 1-based numeric index
    if target.isdigit():
        idx = int(target)
        if 1 <= idx <= len(sheetnames):
            return sheetnames[idx - 1]
        raise SheetNotFound(f"sheet index {idx} out of range (1..{len(sheetnames)})")
    for name in sheetnames:
        if _norm(name) == _norm(target):
            return name
    raise SheetNotFound(f"no sheet named '{sheet}' (have: {', '.join(sheetnames)})")


def _project(
    headers: list[str], rows: list[list[str]], columns: Optional[list[str]]
) -> tuple[list[str], list[list[str]]]:
    if not columns:
        return headers, rows
    idx_by_name = {_norm(h): i for i, h in enumerate(headers)}
    keep: list[int] = []
    out_headers: list[str] = []
    for want in columns:
        i = idx_by_name.get(_norm(str(want)))
        if i is not None:
            keep.append(i)
            out_headers.append(headers[i])
    projected = [[(row[i] if i < len(row) else "") for i in keep] for row in rows]
    return out_headers, projected


def _apply_char_cap(rows: list[list[str]], already_truncated: bool) -> bool:
    total = 0
    for n, row in enumerate(rows):
        total += sum(len(c) for c in row) + len(row)  # +delimiters
        if total > READ_MAX_CHARS:
            del rows[n:]
            return True
    return already_truncated


# --------------------------------------------------------------------------- #
# Public: stream one sheet's rows UNCAPPED (for aggregation)
#
# load_table above materializes the whole grid then windows it — fine for the
# ~200 rows a model can read, wrong for a 200k-row sum. This path yields row by
# row so memory stays flat regardless of file size, and applies NO caps: the
# scan ceiling lives in aggregate.py, where the RESULT is bounded instead.
# --------------------------------------------------------------------------- #
@dataclass
class RowStream:
    sheet_name: str
    headers: list[str]
    rows: Iterator[list[str]]
    all_sheets: list[str]
    total_rows_hint: Optional[int] = None  # data rows; None when unknowable (csv)


@contextmanager
def _csv_stream(path: Path) -> Iterator[Iterable[list[str]]]:
    """Open a CSV and yield a csv.reader, sniffing the delimiter as _csv_grid does."""
    try:
        fh = path.open("r", encoding="utf-8-sig", newline="")
    except OSError as exc:
        raise ReadError(f"could not read CSV: {exc}") from exc
    try:
        sample = fh.read(8192)
        fh.seek(0)
        try:
            dialect = _csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except _csv.Error:
            dialect = _csv.excel  # default to comma
        yield _csv.reader(fh, dialect)
    finally:
        fh.close()


@contextmanager
def open_sheet_rows(path: Path, *, sheet: Optional[str] = None) -> Iterator[RowStream]:
    """Stream ONE sheet: headers plus an iterator over the remaining rows.

    Must be used as a context manager — the workbook/file handle stays open for
    the life of the iterator and is closed on exit, even when the consumer stops
    early (which the aggregator's scan ceiling does by design). Trailing
    all-empty rows are NOT trimmed here, since a stream cannot look ahead; the
    aggregator skips fully blank rows instead.
    """
    path = Path(path)

    if not _is_xlsx(path):
        name = path.stem
        with _csv_stream(path) as reader:
            rows = ([_cell(v) for v in row] for row in reader)
            headers = next(rows, [])
            yield RowStream(
                sheet_name=name,
                headers=headers,
                rows=rows,
                all_sheets=[name],
                total_rows_hint=None,
            )
        return

    wb = _open_xlsx(path)
    try:
        all_sheets = list(wb.sheetnames)
        sheet_name = _resolve_sheet(all_sheets, sheet)
        ws = wb[sheet_name]
        hint = max((ws.max_row or 0) - 1, 0)  # minus the header row
        rows = ([_cell(v) for v in row] for row in ws.iter_rows(values_only=True))
        headers = next(rows, [])
        yield RowStream(
            sheet_name=sheet_name,
            headers=headers,
            rows=rows,
            all_sheets=all_sheets,
            total_rows_hint=hint,
        )
    finally:
        wb.close()


# --------------------------------------------------------------------------- #
# Public: compact summary (for the upload response + the chat attachment note)
# --------------------------------------------------------------------------- #
@dataclass
class Summary:
    kind: str                # "Excel" | "CSV"
    sheets: list[dict]       # [{name, rows, cols, headers}]

    @property
    def total_rows(self) -> int:
        return sum(s["rows"] for s in self.sheets)

    def as_dict(self) -> dict:
        return {"kind": self.kind, "sheets": self.sheets, "total_rows": self.total_rows}

    def text(self) -> str:
        """One-line human/model summary, e.g. 'Excel, 2 sheets, 1240 rows'."""
        n = len(self.sheets)
        if self.kind == "CSV":
            return f"CSV, {self.total_rows} rows"
        sheet_word = "sheet" if n == 1 else "sheets"
        return f"Excel, {n} {sheet_word}, {self.total_rows} rows"


def summarize(path: Path) -> Summary:
    """Structure summary of a file (raises ReadError on a bad file)."""
    path = Path(path)
    infos = inspect_workbook(path)
    kind = "Excel" if _is_xlsx(path) else "CSV"
    sheets = [
        {
            "name": s.sheet_name,
            "rows": s.total_rows,
            "cols": s.total_cols,
            "headers": s.headers,
        }
        for s in infos
    ]
    return Summary(kind=kind, sheets=sheets)
